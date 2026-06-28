from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import unicodedata
import uuid
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, ImageOps

from .field_mapping import (
    DATE_FIELDS,
    EQUIPMENT_COLUMNS,
    EQUIPMENT_END_ROW,
    EQUIPMENT_START_ROW,
    FIELD_ALIASES,
    FIELD_TO_CELL,
    INSUMO_ROWS,
    PERSPECTIVA_ROWS,
    PHOTO_ANCHORS,
    PHOTO_MAX_SIZE,
    SHEET_NAME,
)

ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = ROOT_DIR / "templates" / "relatorio-modelo.xlsx"
DEFAULT_OUTPUT_DIR = ROOT_DIR / "outputs"
PHOTO_BORDER_PX = 3
PHOTO_BORDER_COLOR = (0, 0, 0)
PROPERTY_START_ROW = 18
PROPERTY_TEMPLATE_END_ROW = 21
PROPERTY_INSERT_AT_ROW = 22
PROPERTY_SHIFT_START_ROW = 23
PROPERTY_BASE_CAPACITY = PROPERTY_TEMPLATE_END_ROW - PROPERTY_START_ROW + 1
BENFEITORIA_BASE_BLOCKS = [(27, 29), (30, 31), (32, 33)]
BENFEITORIA_INSERT_AT_ROW = 34
BENFEITORIA_EXTRA_BLOCK_HEIGHT = 3
BENFEITORIA_SIDE = Side(style="thin", color="000000")
CLIENT_NAME_WORD = r"[A-ZÀ-Ú][A-Za-zÀ-ÿ0-9&.'’-]+"
CLIENT_NAME_PARTICLE = r"(?:d[aeo]s?|e)"
CLIENT_NAME_PATTERN = rf"{CLIENT_NAME_WORD}(?:\s+(?:{CLIENT_NAME_PARTICLE}|{CLIENT_NAME_WORD})){{0,7}}"


def normalize_key(value: str) -> str:
    without_accents = unicodedata.normalize("NFKD", str(value))
    without_accents = "".join(ch for ch in without_accents if not unicodedata.combining(ch))
    normalized = re.sub(r"[^a-zA-Z0-9]+", "_", without_accents.lower()).strip("_")
    return normalized


def parse_report_data(raw: str | dict[str, Any]) -> dict[str, Any]:
    if isinstance(raw, dict):
        return normalize_data(raw)

    text = (raw or "").strip()
    if not text:
        return {}

    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return normalize_data(parsed)
    except json.JSONDecodeError:
        pass

    data: dict[str, Any] = {}
    current_key: str | None = None
    loose_lines: list[str] = []

    for original_line in text.splitlines():
        line = original_line.strip()
        if not line:
            if current_key:
                data[current_key] = f"{data[current_key]}\n"
            continue

        match = re.match(r"^([^:=-]{2,80})\s*[:=-]\s*(.+)$", line)
        if match:
            key = FIELD_ALIASES.get(normalize_key(match.group(1)), normalize_key(match.group(1)))
            data[key] = clean_text_value(match.group(2))
            current_key = key
        elif current_key:
            data[current_key] = f"{data[current_key]}\n{line}".strip()
        else:
            loose_lines.append(line)

    if loose_lines:
        data["raw_text"] = "\n".join(loose_lines)

    data.update(parse_narrative_report(text))

    return normalize_data(data)


def parse_narrative_report(text: str) -> dict[str, Any]:
    compact = normalize_spaces(text)
    data: dict[str, Any] = {}

    labeled_fields = {
        "cliente": r"Cliente|Produtor|Mutuário|Mutuario",
        "data_visita": r"Data da visita|Data de visita|Data",
        "cpf_cnpj": r"CPF/CNPJ|CNPJ/CPF|CPF|CNPJ",
        "cidade_uf": r"Município/UF|Municipio/UF|Município|Municipio|Cidade/UF|Cidade",
        "localizacao_1": r"Localização|Localizacao|Endereço/localização|Endereco/localizacao",
        "imovel_nome": r"Nome da propriedade|Nome das propriedades|Nomes das propriedades",
        "tipo_exploracao": r"Tipo de exploração|Tipo de exploracao",
        "atividades_desenvolvidas": r"Atividades desenvolvidas",
        "situacao_produtiva": r"Situação produtiva|Situacao produtiva",
        "area_total_ha": r"Área Total \(ha\)|Area Total \(ha\)",
        "area_pastagens_ha": r"Área de Pastagens \(ha\)|Area de Pastagens \(ha\)",
        "area_cultivo_ha": r"Área de Cultivo \(ha\)|Area de Cultivo \(ha\)",
        "atividade_principal": r"Atividade principal desenvolvida",
        "principais_culturas": r"Principais culturas",
    }
    label_patterns = list(labeled_fields.values())
    section_patterns = [
        r"\d+\.\s*TIPO",
        r"\d+\.\s*DESCRIÇÃO|DESCRICAO",
        r"INVESTIMENTOS EM ANDAMENTO",
        r"OUTROS COMENTÁRIOS|OUTROS COMENTARIOS",
        r"CONCLUSÃO|CONCLUSAO",
        r"FRASES DIRETAS",
        r"Dados\s+de\s+(?:Área|Area)\s+e\s+(?:Exploração|Exploracao)",
    ]
    stop_pattern = "|".join([rf"(?:{pattern})\s*:" for pattern in label_patterns] + section_patterns)

    for field, label_pattern in labeled_fields.items():
        value = extract_after_label(compact, label_pattern, stop_pattern)
        if value:
            data[field] = value
    if not data.get("data_visita"):
        visit_date = extract_visit_date(compact)
        if visit_date:
            data["data_visita"] = visit_date

    explicit_property_summary = data.get("imovel_nome")
    structured_properties = parse_property_area_sections(text)
    if structured_properties:
        data["imoveis"] = structured_properties
        if explicit_property_summary:
            data["imovel_resumo"] = explicit_property_summary
        data["imovel_nome"] = format_property_names([str(item["nome"]) for item in structured_properties if item.get("nome")])
        totals = aggregate_property_totals(structured_properties)
        detail_only = all(
            item.get("atividade_principal") in (None, "") and item.get("principais_culturas") in (None, "")
            for item in structured_properties
        )
        for key, value in totals.items():
            if value in (None, ""):
                continue
            if detail_only:
                data.setdefault(key, value)
            else:
                data[key] = value
    else:
        property_names = extract_property_names(compact, data.get("imovel_nome"), stop_pattern)
        if len(property_names) > 1:
            data["imoveis"] = [{"nome": name} for name in property_names]

    property_names = property_names_from_value(data.get("imoveis")) or extract_property_names(compact, data.get("imovel_nome"), stop_pattern)
    if property_names and not structured_properties:
        if len(property_names) > 1:
            data["imoveis"] = [{"nome": name} for name in property_names]
            data["imovel_nome"] = format_property_names(property_names)
        elif not data.get("imovel_nome"):
            data["imovel_nome"] = property_names[0]

    for area_field in ("area_total_ha", "area_pastagens_ha", "area_cultivo_ha"):
        if data.get(area_field):
            data[area_field] = parse_decimal_pt(data[area_field])

    if data.get("area_total_ha") not in (None, "") and data.get("area_pastagens_ha") not in (None, "") and not data.get("area_cultivo_ha"):
        data["area_cultivo_ha"] = round(float(data["area_total_ha"]) - float(data["area_pastagens_ha"]), 2)

    if not data.get("atividade_principal") and data.get("atividades_desenvolvidas"):
        data["atividade_principal"] = data["atividades_desenvolvidas"]

    benfeitorias_geral = None
    benfeitorias = extract_section(
        compact,
        r"\d+\.\s*TIPO\s*\(Benfeitorias e Infraestrutura\)",
        [r"\d+\.\s*DESCRIÇÃO", r"\d+\.\s*DESCRICAO", r"INVESTIMENTOS EM ANDAMENTO"],
    )
    if benfeitorias:
        benfeitorias_propriedades, benfeitorias_geral = split_benfeitorias_property_scope(benfeitorias)
        data["benfeitorias_descricao"] = benfeitorias_propriedades or benfeitorias
        data.setdefault("benfeitorias_conservacao", "BOM")
        data.setdefault(
            "benfeitorias_observacoes",
            "Estrutura compativel com a escala produtiva informada.",
        )

    equipamentos = extract_section(
        compact,
        r"\d+\.\s*DESCRIÇÃO\s*\(Máquinas, Equipamentos e Implementos\)|\d+\.\s*DESCRICAO\s*\(Maquinas, Equipamentos e Implementos\)",
        [r"INVESTIMENTOS EM ANDAMENTO"],
    )
    if equipamentos:
        data["equipamentos"] = parse_equipment_section(equipamentos)

    investimentos = extract_section(
        compact,
        r"INVESTIMENTOS EM ANDAMENTO\s*\(Comentários\)|INVESTIMENTOS EM ANDAMENTO\s*\(Comentarios\)",
        [r"OUTROS COMENTÁRIOS", r"OUTROS COMENTARIOS", r"CONCLUSÃO", r"CONCLUSAO"],
    )
    if investimentos:
        data["investimentos_comentarios"] = investimentos

    outros = extract_section(
        compact,
        r"OUTROS COMENTÁRIOS|OUTROS COMENTARIOS",
        [r"CONCLUSÃO", r"CONCLUSAO", r"FRASES DIRETAS"],
    )
    if outros:
        data["outros_comentarios"] = outros
    if benfeitorias_geral:
        data["outros_comentarios"] = append_comment(data.get("outros_comentarios"), benfeitorias_geral)

    conclusao = extract_section(
        compact,
        r"CONCLUSÃO|CONCLUSAO",
        [r"FRASES DIRETAS"],
    )
    if conclusao:
        data["conclusao"] = conclusao

    frase_direta = extract_section(
        compact,
        r"FRASES? DIRETAS?\s*(?:\(PADRÃO DE MATRÍCULA/VISUALIZAÇÃO\)|\(PADRAO DE MATRICULA/VISUALIZACAO\))?",
        [],
    )
    if frase_direta:
        data["insumos_comentarios"] = frase_direta

    city = extract_city(compact)
    if city:
        data.setdefault("cidade_uf", city)
        data.setdefault("localizacao_1", city)

    area_alqueires = extract_area_alqueires(compact)
    if area_alqueires:
        data["area_alqueires"] = area_alqueires

    client = extract_client_name(compact)
    if client and not data.get("cliente"):
        data["cliente"] = client

    rebanho = extract_rebanho(compact)
    if rebanho:
        data["rebanho"] = rebanho

    enrich_location_summary(data)

    return data


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def clean_text_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    return value.strip().strip('"').strip("'").strip()


def extract_after_label(text: str, label_pattern: str, stop_pattern: str) -> str | None:
    match = re.search(
        rf"(?:{label_pattern})\s*:\s*(.*?)(?=\s*(?:{stop_pattern})|$)",
        text,
        flags=re.IGNORECASE,
    )
    return clean_text_value(match.group(1)) if match else None


def extract_section(text: str, start_pattern: str, end_patterns: list[str]) -> str | None:
    end_pattern = "|".join(end_patterns)
    lookahead = rf"(?=\s*(?:{end_pattern})|$)" if end_pattern else r"(?=$)"
    match = re.search(
        rf"(?:{start_pattern})\s*(.*?){lookahead}",
        text,
        flags=re.IGNORECASE,
    )
    return clean_text_value(match.group(1)) if match else None


def extract_visit_date(text: str) -> str | None:
    patterns = [
        r"(?:Data\s+da\s+visita|Data\s+de\s+visita|Data)\s*:?\s*(\d{1,2}[\/.-]\d{1,2}[\/.-]\d{2,4})",
        r"(?:visita\s+(?:realizada|efetuada)\s+em)\s*:?\s*(\d{1,2}[\/.-]\d{1,2}[\/.-]\d{2,4})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return None


def append_comment(existing: Any, addition: str) -> str:
    current = normalize_spaces(str(existing or ""))
    extra = normalize_spaces(addition)
    if not current:
        return extra
    if normalize_key(extra) in normalize_key(current):
        return current
    return f"{current}\n{extra}"


def parse_decimal_pt(value: Any) -> float | Any:
    if isinstance(value, (int, float)):
        return value

    match = re.search(r"\d[\d.,]*", str(value))
    if not match:
        return value

    number = match.group(0)
    if "," in number and "." in number:
        decimal_separator = "," if number.rfind(",") > number.rfind(".") else "."
        thousand_separator = "." if decimal_separator == "," else ","
        number = number.replace(thousand_separator, "").replace(decimal_separator, ".")
    elif "," in number:
        number = number.replace(".", "").replace(",", ".")
    elif number.count(".") > 1:
        parts = number.split(".")
        number = "".join(parts[:-1]) + "." + parts[-1]
    elif "." in number:
        integer, decimal = number.split(".", 1)
        if len(decimal) == 3 and len(integer) <= 3:
            number = integer + decimal

    return float(number)


def extract_city(text: str) -> str | None:
    patterns = [
        r"município de\s+([A-ZÀ-ÚA-Za-zà-úÇçãõíóéâêôü\s]+-[A-Z]{2})",
        r"municipio de\s+([A-ZÀ-ÚA-Za-zà-úÇçãõíóéâêôü\s]+-[A-Z]{2})",
        r"localizad[oa]\s+em\s+([A-ZÀ-ÚA-Za-zà-úÇçãõíóéâêôü\s]+-[A-Z]{2})",
        r"\bem\s+([A-ZÀ-ÚA-Za-zà-úÇçãõíóéâêôü\s]+-[A-Z]{2})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return normalize_spaces(match.group(1))
    return None


def extract_area_alqueires(text: str) -> str | None:
    match = re.search(r"\((\d+(?:[.,]\d+)?)\s*alqueires?\)", text, flags=re.IGNORECASE)
    return match.group(1) if match else None


def extract_client_name(text: str) -> str | None:
    patterns = [
        rf"\b(?i:produtor(?:a)?(?:\s+rural)?)\s+({CLIENT_NAME_PATTERN})(?=\s+(?i:em|no|na|est[aá]|desenvolve|conduz|comanda|consolida|det[eé]m|mant[eé]m|mantem|possui|opera|localizad[oa])\b|[,.;]|$)",
        rf"\b(?i:mutu[aá]rio(?:a)?)\s+({CLIENT_NAME_PATTERN})(?=\s+(?i:em|no|na|est[aá]|desenvolve|conduz|comanda|consolida|det[eé]m|mant[eé]m|mantem|possui|opera|localizad[oa])\b|[,.;]|$)",
        rf"\b(?i:cliente)\s+({CLIENT_NAME_PATTERN})(?=\s+(?i:em|no|na|est[aá]|desenvolve|conduz|comanda|consolida|det[eé]m|mant[eé]m|mantem|possui|opera|localizad[oa])\b|[,.;]|$)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            candidate = sanitize_client_name(match.group(1))
            if candidate:
                return candidate
    return sanitize_client_name(text)


def sanitize_client_name(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = normalize_spaces(str(value))
    if not text:
        return None

    direct_match = re.match(
        rf"^({CLIENT_NAME_PATTERN})(?=\s+(?i:em|no|na|est[aá]|desenvolve|conduz|comanda|consolida|det[eé]m|mant[eé]m|mantem|possui|opera|localizad[oa])\b|[,.;]|$)",
        text,
    )
    candidate = direct_match.group(1) if direct_match else text
    candidate = clean_text_value(candidate)
    candidate = re.sub(r"\s+(?:CPF|CNPJ)\b.*$", "", candidate, flags=re.IGNORECASE).strip()
    candidate = candidate.strip(" -;,.")

    if not looks_like_client_name(candidate):
        return None
    return candidate


def looks_like_client_name(value: str) -> bool:
    text = normalize_spaces(value)
    if not text or len(text) > 80:
        return False
    words = text.split()
    if len(words) > 8:
        return False
    rejected_terms = {
        "fazenda",
        "sitio",
        "chacara",
        "propriedade",
        "imovel",
        "municipio",
        "hectare",
        "hectares",
        "pecuaria",
        "agricola",
        "operacao",
        "infraestrutura",
        "benfeitoria",
        "pastagem",
        "pastagens",
        "plantel",
        "cabecas",
        "complexo",
    }
    normalized = normalize_key(text)
    return not any(term in normalized for term in rejected_terms)


def extract_rebanho(text: str) -> str | None:
    patterns = [
        r"plantel total informado\s+(?:é\s+de|e\s+de|de)\s+(.+?)(?=\.|$)",
        r"rebanho consolidado de\s+(.+?)(?=\.|$)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return clean_text_value(match.group(1))
    return None


def parse_property_area_sections(text: str) -> list[dict[str, Any]]:
    compact = normalize_spaces(text)
    section = extract_section(
        compact,
        r"Dados\s+de\s+(?:Área|Ãrea|Area)\s+e\s+(?:Exploração|ExploraÃ§Ã£o|Exploracao)\s+por\s+Propriedade",
        [r"\d+\.\s*TIPO", r"\d+\.\s*DESCRIÇÃO", r"\d+\.\s*DESCRIÃ‡ÃƒO", r"\d+\.\s*DESCRICAO"],
    )
    properties = parse_compact_property_area_section(section or "")
    if properties:
        return properties

    detail_properties = parse_unit_detail_sections(compact)
    if detail_properties:
        return detail_properties

    lines = [line.strip() for line in text.splitlines()]
    start_index = None
    for index, line in enumerate(lines):
        normalized = normalize_key(line)
        if "dados_de_area" in normalized and "propriedade" in normalized:
            start_index = index + 1
            break

    if start_index is None:
        return []

    properties: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    for raw_line in lines[start_index:]:
        line = raw_line.strip()
        if not line:
            continue

        normalized = normalize_key(line)
        if re.match(r"^\d+\.\s*(?:TIPO|DESCRICAO|DESCRI)", line, flags=re.IGNORECASE) or normalized.startswith("tipo_benfeitorias"):
            break

        if is_property_heading(line):
            current = {"nome": clean_property_name(line)}
            properties.append(current)
            continue

        if current is None:
            continue

        match = re.match(r"^([^:]{2,90})\s*:\s*(.+)$", line)
        if not match:
            continue

        key = normalize_key(match.group(1))
        value = clean_text_value(match.group(2))
        if key.startswith("area_total"):
            current["area_total_ha"] = parse_decimal_pt(value)
            alqueires = extract_area_alqueires(str(value))
            if alqueires:
                current["area_alqueires"] = alqueires
        elif "pastagens" in key:
            current["area_pastagens_ha"] = parse_decimal_pt(value)
        elif "cultivo" in key:
            current["area_cultivo_ha"] = parse_decimal_pt(value)
        elif "atividade_principal" in key or key.startswith("atividade"):
            current["atividade_principal"] = value
        elif "principais_culturas" in key or key == "culturas":
            current["principais_culturas"] = value

    return [item for item in properties if item.get("nome")]


def parse_compact_property_area_section(section: str) -> list[dict[str, Any]]:
    if not section:
        return []

    area_label = r"(?:Área|Ãrea|Area)\s+Total\s*\(ha\)\s*:"
    pasture_label = r"(?:Área|Ãrea|Area)\s+de\s+Pastagens\s*\(ha\)\s*:"
    crop_label = r"(?:Área|Ãrea|Area)\s+de\s+Cultivo\s*\(ha\)\s*:"
    activity_label = r"Atividade\s+principal\s+desenvolvida\s*:"
    cultures_label = r"Principais\s+culturas\s*:"
    property_heading = r"(?:Fazenda|S[ií]tio|Sitio|Ch[aá]cara|Chacara|Grupo|Unidade|Propriedade)\s+"
    next_property = rf"(?=\s*{property_heading}.{{2,180}}?{area_label}|$)"
    pattern = re.compile(
        rf"\s*(?P<name>{property_heading}.+?)\s*{area_label}\s*(?P<total>.*?)\s*"
        rf"{pasture_label}\s*(?P<pasture>.*?)\s*"
        rf"{crop_label}\s*(?P<crop>.*?)\s*"
        rf"{activity_label}\s*(?P<activity>.*?)\s*"
        rf"{cultures_label}\s*(?P<cultures>.*?){next_property}",
        flags=re.IGNORECASE,
    )

    properties: list[dict[str, Any]] = []
    for match in pattern.finditer(section):
        name = clean_property_name(match.group("name"))
        if not name:
            continue
        item: dict[str, Any] = {
            "nome": name,
            "area_total_ha": parse_decimal_pt(match.group("total")),
            "area_pastagens_ha": parse_decimal_pt(match.group("pasture")),
            "area_cultivo_ha": parse_decimal_pt(match.group("crop")),
            "atividade_principal": clean_property_field(match.group("activity")),
            "principais_culturas": clean_property_field(match.group("cultures")),
        }
        alqueires = extract_area_alqueires(match.group("total"))
        if alqueires:
            item["area_alqueires"] = parse_decimal_pt(alqueires)
        properties.append(item)

    return properties


def parse_unit_detail_sections(text: str) -> list[dict[str, Any]]:
    section = extract_section(
        text,
        r"Detalhamento\s+por\s+Unidade\s*\(ha\)",
        [r"\d+\.\s*TIPO", r"\d+\.\s*DESCRIÇÃO", r"\d+\.\s*DESCRIÃ‡ÃƒO", r"\d+\.\s*DESCRICAO"],
    )
    if not section:
        return []

    pattern = re.compile(
        r"(?P<name>(?:Fazenda|S[iÃ­]tio|Sitio|Ch[aÃ¡]cara|Chacara|Grupo|Unidade|Propriedade)\s+.+?)\s*:\s*"
        r"(?P<total>\d[\d.,]*)\s*ha\s+brutos?\s*/\s*"
        r"(?P<pasture>\d[\d.,]*)\s*ha\s+de\s+pastagens",
        flags=re.IGNORECASE,
    )
    properties: list[dict[str, Any]] = []
    for match in pattern.finditer(section):
        total = parse_decimal_pt(match.group("total"))
        pasture = parse_decimal_pt(match.group("pasture"))
        item: dict[str, Any] = {
            "nome": clean_property_name(match.group("name")),
            "area_total_ha": total,
            "area_pastagens_ha": pasture,
        }
        if isinstance(total, (int, float)) and isinstance(pasture, (int, float)):
            item["area_cultivo_ha"] = round(float(total) - float(pasture), 2)
        properties.append(item)
    return properties


def clean_property_field(value: Any) -> str:
    text = str(clean_text_value(value) or "")
    text = re.sub(r"\s+\d+\.\s*(?:TIPO|DESCRIÃ‡ÃƒO|DESCRICAO)\b.*$", "", text, flags=re.IGNORECASE)
    return text.strip(" .;:-")


def is_property_heading(value: str) -> bool:
    if ":" in value or len(value) > 90:
        return False
    return starts_with_property_type(value)


def aggregate_property_totals(properties: list[dict[str, Any]]) -> dict[str, Any]:
    totals: dict[str, Any] = {}
    numeric_fields = ("area_total_ha", "area_pastagens_ha", "area_cultivo_ha", "area_financiada_bb_ha", "area_financiada_outros_ha")
    for field in numeric_fields:
        values = [item.get(field) for item in properties if isinstance(item.get(field), (int, float))]
        if values:
            totals[field] = round(sum(float(value) for value in values), 2)

    first_activity = next((item.get("atividade_principal") for item in properties if item.get("atividade_principal")), None)
    first_cultures = next((item.get("principais_culturas") for item in properties if item.get("principais_culturas")), None)
    if first_activity and all(item.get("atividade_principal") in (None, "", first_activity) for item in properties):
        totals["atividade_principal"] = first_activity
    if first_cultures and all(item.get("principais_culturas") in (None, "", first_cultures) for item in properties):
        totals["principais_culturas"] = first_cultures
    return totals


def extract_property_names(text: str, explicit_value: Any = None, stop_pattern: str | None = None) -> list[str]:
    names = property_names_from_value(explicit_value)
    patterns = [
        r"(?:im[oó]veis?\s+rurais?|propriedades?|fazendas?|unidades?)\s+(?:denominad[oa]s?|chamad[oa]s?)\s+(.+?)(?=\s+(?:localizad|situad|compreend|abrang|totaliz|possu|no\s+municipio|na\s+regiao)\b|[.;]|$)",
        r"(?:possui|det[eé]m|explora|administra|opera|trabalha)\s+(?:as?|os?)?\s*(?:im[oó]veis?\s+rurais?|propriedades?|fazendas?)\s+(.+?)(?=\s+(?:localizad|situad|compreend|abrang|totaliz|no\s+municipio|na\s+regiao)\b|[.;]|$)",
    ]
    if stop_pattern:
        value = extract_after_label(
            text,
            r"Nome(?:s)?\s+d(?:a|as)\s+propriedade(?:s)?|Propriedade(?:s)?|Fazenda(?:s)?|Im[oó]ve(?:l|is)(?:\s+rurais?)?",
            stop_pattern,
        )
        names.extend(split_property_names(value))

    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            names.extend(split_property_names(match.group(1)))

    return dedupe_property_names(names)


def property_names_from_value(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        names: list[str] = []
        for item in value:
            if isinstance(item, dict):
                names.extend(split_property_names(item.get("nome", item.get("name", ""))))
            else:
                names.extend(split_property_names(item))
        return names
    return split_property_names(value)


def split_property_names(value: Any) -> list[str]:
    if value in (None, ""):
        return []

    raw = normalize_spaces(str(value))
    if not raw:
        return []

    parts: list[str] = []
    for chunk in split_outside_parentheses(raw, ";,/\n"):
        parts.extend(split_conjunction_property_names(chunk))

    return [name for name in (clean_property_name(part) for part in parts) if looks_like_property_name(name)]


def split_outside_parentheses(value: str, separators: str) -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    depth = 0
    for char in value:
        if char == "(":
            depth += 1
        elif char == ")" and depth:
            depth -= 1
        if char in separators and depth == 0:
            part = "".join(current).strip()
            if part:
                parts.append(part)
            current = []
            continue
        current.append(char)
    part = "".join(current).strip()
    if part:
        parts.append(part)
    return parts


def split_conjunction_property_names(value: str) -> list[str]:
    pieces = re.split(r"\s+e\s+(?=(?:Fazenda|Sitio|S[ií]tio|Chacara|Ch[aá]cara|Estancia|Est[aâ]ncia|Rancho|Gleba|Granja|Retiro|Lote|Im[oó]vel|Propriedade)\b)", value, flags=re.IGNORECASE)
    if len(pieces) > 1:
        return pieces

    if re.search(r"\b(Fazenda|Sitio|S[ií]tio|Chacara|Ch[aá]cara|Estancia|Est[aâ]ncia|Rancho|Gleba|Granja|Retiro)\b", value, flags=re.IGNORECASE):
        implicit = re.split(r"\s+e\s+", value, flags=re.IGNORECASE)
        if len(implicit) > 1:
            prefix_match = re.match(r"\s*(Fazenda|Sitio|S[ií]tio|Chacara|Ch[aá]cara|Estancia|Est[aâ]ncia|Rancho|Gleba|Granja|Retiro)\b", implicit[0], flags=re.IGNORECASE)
            if prefix_match:
                prefix = prefix_match.group(1)
                return [implicit[0], *[part if starts_with_property_type(part) else f"{prefix} {part}" for part in implicit[1:]]]

    return [value]


def clean_property_name(value: str) -> str:
    cleaned = clean_text_value(value)
    cleaned = re.sub(r"\s+\d+\.\s*(?:TIPO|DESCRIÇÃO|DESCRIÃ‡ÃƒO|DESCRICAO)\b.*$", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(
        r"\s+(?:Tipo\s+de\s+exploração|Tipo\s+de\s+exploraÃ§Ã£o|Tipo\s+de\s+exploracao|Atividades\s+desenvolvidas|Situação\s+produtiva|SituaÃ§Ã£o\s+produtiva|Situacao\s+produtiva|Dados\s+de\s+(?:Área|Ãrea|Area))\b.*$",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"^(?:as?|os?|das?|dos?|e|seguintes?)\s+", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+(?:localizad[oa]s?|situad[oa]s?|compreend(?:e|em)|abrang(?:e|em)|totaliz(?:a|am)|possu(?:i|em))\b.*$", "", cleaned, flags=re.IGNORECASE)
    return cleaned.strip(" .;:-")


def looks_like_property_name(value: str) -> bool:
    if not value:
        return False
    normalized = normalize_key(value)
    property_terms = {"fazenda", "sitio", "chacara", "estancia", "rancho", "gleba", "granja", "retiro", "lote", "imovel", "propriedade"}
    return any(term in normalized.split("_") for term in property_terms)


def starts_with_property_type(value: str) -> bool:
    return bool(re.match(r"\s*(Fazenda|Sitio|S[ií]tio|Chacara|Ch[aá]cara|Estancia|Est[aâ]ncia|Rancho|Gleba|Granja|Retiro|Lote|Im[oó]vel|Propriedade)\b", value, flags=re.IGNORECASE))


def dedupe_property_names(names: list[str]) -> list[str]:
    deduped: list[str] = []
    seen: set[str] = set()
    for name in names:
        key = normalize_key(name)
        if not key or key in seen:
            continue
        seen.add(key)
        deduped.append(name)
    return deduped


def format_property_names(names: list[str]) -> str:
    return "; ".join(name for name in names if str(name).strip())


def parse_equipment_section(section: str) -> list[dict[str, Any]]:
    if re.search(
        r"não dispõe|nao dispoe|não informado|nao informado|não detalhou|nao detalhou",
        section,
        flags=re.IGNORECASE,
    ):
        return [
            {
                "descricao": "Não informado",
                "fabricante": "-",
                "modelo": "-",
                "estado": "BOM",
                "financiado_bb": "NÃO",
                "financiado_outros": "NÃO",
                "segurado": "NÃO",
                "gravame": "NÃO",
                "outras_informacoes": "Não há frota mecanizada própria ou sistema de irrigação declarado.",
            }
        ]
    description = extract_after_label(section, r"Descrição|Descricao", r"(?:Fabricante|Modelo)\s*:|$")
    manufacturer = extract_after_label(section, r"Fabricante", r"(?:Modelo|Estado)\s*:|$")
    model = extract_after_label(section, r"Modelo", r"(?:Estado|Conservação|Conservacao)\s*:|$")
    if description:
        return [
            {
                "descricao": description,
                "fabricante": manufacturer or "-",
                "modelo": model or "-",
                "estado": "BOM",
                "financiado_bb": "NÃO",
                "financiado_outros": "NÃO",
                "segurado": "NÃO",
                "gravame": "NÃO",
                "outras_informacoes": "",
            }
        ]
    cleaned = clean_equipment_description(section)
    if cleaned:
        items = [part.strip(" .;:-") for part in re.split(r"\s*;\s*", cleaned) if part.strip(" .;:-")]
        if not items:
            items = [cleaned]
        return [
            {
                "descricao": item,
                "fabricante": "-",
                "modelo": "-",
                "estado": "BOM",
                "financiado_bb": "NÃO",
                "financiado_outros": "NÃO",
                "segurado": "NÃO",
                "gravame": "NÃO",
                "outras_informacoes": "",
            }
            for item in items
        ]
    return []


def clean_equipment_description(value: Any) -> str:
    text = normalize_spaces(str(value or ""))
    text = re.sub(r"^(?:Descri[cç][aã]o|Descricao)\s*:\s*", "", text, flags=re.IGNORECASE)
    return text.strip(" .;:-")


def normalize_date_text(value: Any) -> Any:
    match = re.search(r"\d{1,2}[\/.-]\d{1,2}[\/.-]\d{2,4}", str(value))
    return match.group(0) if match else value


def normalize_data(data: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}

    for key, value in data.items():
        normalized_key = FIELD_ALIASES.get(normalize_key(key), normalize_key(key))

        if isinstance(value, dict):
            normalized[normalized_key] = normalize_nested_dict(value)
        elif isinstance(value, list):
            normalized[normalized_key] = value
        else:
            normalized[normalized_key] = value

    if normalized.get("cliente"):
        cleaned_client = sanitize_client_name(normalized.get("cliente"))
        if cleaned_client:
            normalized["cliente"] = cleaned_client
        elif len(str(normalized["cliente"])) > 80:
            normalized.pop("cliente", None)

    if normalized.get("data_visita"):
        normalized["data_visita"] = normalize_date_text(normalized["data_visita"])
        normalized.setdefault("data_assinatura", normalized["data_visita"])
        normalized.setdefault("data_administracao", normalized["data_visita"])

    if isinstance(normalized.get("equipamentos"), str):
        normalized["equipamentos"] = parse_equipment_section(str(normalized["equipamentos"]))

    property_items = normalize_property_items(normalized.get("imoveis"))
    property_names = property_names_from_value(property_items)
    if property_items:
        normalized["imoveis"] = property_items
        normalized.setdefault("imovel_nome", format_property_names(property_names))
        totals = aggregate_property_totals(property_items)
        for key, value in totals.items():
            normalized.setdefault(key, value)
    else:
        property_names = property_names_from_value(normalized.get("imovel_nome"))
        if len(property_names) > 1:
            normalized["imoveis"] = [{"nome": name} for name in property_names]
            normalized["imovel_nome"] = format_property_names(property_names)

    return normalized


def normalize_nested_dict(value: dict[str, Any]) -> dict[str, Any]:
    return {normalize_key(k): v for k, v in value.items()}


def normalize_property_items(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []

    items: list[dict[str, Any]] = []
    for item in value:
        if isinstance(item, dict):
            normalized_item: dict[str, Any] = {}
            for key, raw_value in item.items():
                normalized_key = FIELD_ALIASES.get(normalize_key(key), normalize_key(key))
                normalized_item[normalized_key] = parse_decimal_pt(raw_value) if normalized_key.startswith("area_") and raw_value not in (None, "") else raw_value
            if normalized_item.get("nome") or normalized_item.get("imovel_nome"):
                normalized_item["nome"] = clean_property_name(str(normalized_item.get("nome", normalized_item.get("imovel_nome"))))
                items.append(normalized_item)
        else:
            for name in split_property_names(item):
                items.append({"nome": name})

    return items


def generate_report(
    data_source: str | dict[str, Any],
    photo_paths: list[str | Path] | None = None,
    output_path: str | Path | None = None,
    template_path: str | Path = DEFAULT_TEMPLATE,
) -> Path:
    data = parse_report_data(data_source)
    template = Path(template_path)

    if not template.exists():
        raise FileNotFoundError(f"Modelo nao encontrado: {template}")

    run_id = datetime.now().strftime("%Y%m%d-%H%M%S") + "-" + uuid.uuid4().hex[:8]
    output = Path(output_path) if output_path else DEFAULT_OUTPUT_DIR / f"relatorio-gerado-{run_id}.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(template)
    worksheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active

    row_offset = prepare_property_rows(worksheet, data)
    clear_variable_model_values(worksheet, row_offset)
    benfeitoria_blocks = split_benfeitoria_blocks(data.get("benfeitorias_descricao"))
    benfeitoria_offset = prepare_benfeitoria_rows(worksheet, benfeitoria_blocks, row_offset)
    below_benfeitoria_offset = row_offset + benfeitoria_offset
    apply_fields(worksheet, data, row_offset, below_benfeitoria_offset)
    apply_benfeitoria_blocks(worksheet, data, benfeitoria_blocks, row_offset)
    apply_property_rows(worksheet, data)
    apply_equipment(worksheet, data.get("equipamentos", []), below_benfeitoria_offset)
    # Insumos: uniao de duas fontes que NUNCA inventam.
    #  - rede de seguranca deterministica: marca so o que esta literalmente
    #    escrito (corrego/tanque -> agua; placa solar -> energia), cobrindo
    #    casos em que a IA, por ser nao-deterministica, esquece de marcar;
    #  - decisao da IA: refina com logica de contexto.
    insumos = infer_insumos(data)
    ai_insumos = data.get("insumos")
    if isinstance(ai_insumos, dict):
        insumos.update(ai_insumos)
    apply_insumos(worksheet, insumos, below_benfeitoria_offset)
    apply_perspectivas(worksheet, data.get("perspectivas", {}), below_benfeitoria_offset)
    polish_written_ranges(worksheet, row_offset, below_benfeitoria_offset)
    adjust_dynamic_row_heights(worksheet)
    apply_photos(worksheet, photo_paths or [], output.parent / f"{output.stem}-images", below_benfeitoria_offset)

    workbook.save(output)
    return output


def prepare_property_rows(worksheet: Worksheet, data: dict[str, Any]) -> int:
    properties = normalize_property_items(data.get("imoveis"))
    extra_rows = max(0, len(properties) - PROPERTY_BASE_CAPACITY)
    if not extra_rows:
        return 0

    shifted_merges = collect_shifted_merged_ranges(worksheet, PROPERTY_INSERT_AT_ROW, extra_rows)
    worksheet.insert_rows(PROPERTY_INSERT_AT_ROW, extra_rows)
    for merge_range in shifted_merges:
        worksheet.merge_cells(merge_range)
    for row in range(PROPERTY_INSERT_AT_ROW, PROPERTY_INSERT_AT_ROW + extra_rows):
        copy_row_style(worksheet, PROPERTY_TEMPLATE_END_ROW, row)
        merge_property_row(worksheet, row)
    return extra_rows


def split_benfeitoria_blocks(value: Any) -> list[str]:
    if value in (None, ""):
        return []

    text = normalize_spaces(str(value))
    if not text:
        return []

    start_pattern = r"\b(?:Na|No)\s+(?:Fazenda|Unidade|Grupo|S[ií]tio|Ch[aá]cara)\b"
    matches = list(re.finditer(start_pattern, text))
    if not matches:
        return split_single_benfeitoria_block(text)

    blocks: list[str] = []
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        block = text[match.start() : end].strip()
        if block:
            if len(matches) == 1:
                blocks.extend(split_single_benfeitoria_block(block))
            else:
                blocks.append(block)
    return blocks


def split_single_benfeitoria_block(text: str) -> list[str]:
    text = normalize_spaces(text)
    if not should_split_single_benfeitoria_block(text):
        return [text] if text else []

    sentences = split_sentences_pt(text)
    if len(sentences) > 1:
        return group_text_blocks(sentences, target_length=320)

    listed_blocks = split_listed_benfeitoria_items(text)
    return listed_blocks or [text]


def should_split_single_benfeitoria_block(text: str) -> bool:
    normalized = normalize_key(text)
    if any(
        marker in normalized
        for marker in (
            "nao_foram_detalhadas",
            "nao_foram_detalhados",
            "nao_foram_discriminadas",
            "nao_foram_discriminados",
            "nao_foram_informadas",
            "nao_foram_informados",
            "nao_dispoe",
            "nao_detalhou",
        )
    ):
        return False

    concrete_terms = (
        "curral",
        "piquete",
        "cocho",
        "bebedouro",
        "galpao",
        "barracao",
        "casa",
        "cerca",
        "tanque",
        "poco",
        "represa",
        "silo",
        "trincheira",
        "confinamento",
        "energia",
        "placa_solar",
        "pastagem",
    )
    hits = {term for term in concrete_terms if term in normalized}
    return len(hits) >= 3 or (len(hits) >= 2 and len(text) >= 320)


def split_sentences_pt(text: str) -> list[str]:
    return [
        sentence.strip()
        for sentence in re.split(r"(?<=[.!?])\s+(?=[A-Z0-9])", text)
        if sentence.strip()
    ]


def group_text_blocks(parts: list[str], target_length: int = 320) -> list[str]:
    blocks: list[str] = []
    current = ""
    for part in parts:
        candidate = f"{current} {part}".strip() if current else part
        if current and len(candidate) > target_length:
            blocks.append(current)
            current = part
            continue
        current = candidate
    if current:
        blocks.append(current)
    return blocks


def split_listed_benfeitoria_items(text: str) -> list[str]:
    match = re.match(
        r"^(?P<prefix>.*?\b(?:foi informado|foi informada|foram informados|foram informadas|conta com|disp[oÃµ]e de|possui|inclui)\b)\s+(?P<tail>.+)$",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        return []

    tail = match.group("tail").strip()
    tail = re.sub(
        r"\s+e\s+(?=(?:curral|piquete|cocho|bebedouro|galp[aÃ£]o|barrac[aÃ£]o|casa|cerca|tanque|po[cÃ§]o|represa|silo|trincheira|confinamento|pastagem)\b)",
        ", ",
        tail,
        flags=re.IGNORECASE,
    )
    items = [item.strip(" .;") for item in tail.split(",") if item.strip(" .;")]
    if len(items) < 2:
        return []

    blocks = [f"{match.group('prefix').strip()} {items[0]}."]
    blocks.extend(group_text_blocks([f"{item}." for item in items[1:]], target_length=240))
    return blocks


def split_benfeitorias_property_scope(value: Any) -> tuple[str, str | None]:
    text = normalize_spaces(str(value or ""))
    if not text:
        return "", None

    start_pattern = r"\b(?:Na|No)\s+(?:Fazenda|Unidade|Grupo|S[iÃ­]tio|Ch[aÃ¡]cara)\b"
    matches = list(re.finditer(start_pattern, text))
    if not matches:
        return text, None

    general = text[: matches[0].start()].strip(" .;:-")
    property_text = text[matches[0].start() :].strip()
    return property_text, general or None


def prepare_benfeitoria_rows(worksheet: Worksheet, blocks: list[str], property_offset: int = 0) -> int:
    extra_blocks = max(0, len(blocks) - len(BENFEITORIA_BASE_BLOCKS))
    extra_rows = extra_blocks * BENFEITORIA_EXTRA_BLOCK_HEIGHT
    if not extra_rows:
        return 0

    insert_at = BENFEITORIA_INSERT_AT_ROW + property_offset
    shifted_merges = collect_shifted_merged_ranges(worksheet, insert_at, extra_rows)
    worksheet.insert_rows(insert_at, extra_rows)
    for merge_range in shifted_merges:
        worksheet.merge_cells(merge_range)

    source_rows = [BENFEITORIA_BASE_BLOCKS[-1][0] + property_offset, BENFEITORIA_BASE_BLOCKS[-1][1] + property_offset]
    for index, row in enumerate(range(insert_at, insert_at + extra_rows)):
        copy_row_style(worksheet, source_rows[index % len(source_rows)], row)
        worksheet.row_dimensions[row].height = 24
    return extra_rows


def collect_shifted_merged_ranges(worksheet: Worksheet, insert_at: int, amount: int) -> list[str]:
    shifted: list[str] = []
    for merged_range in list(worksheet.merged_cells.ranges):
        range_string = str(merged_range)
        min_col, min_row, max_col, max_row = range_boundaries(range_string)
        if min_row >= insert_at:
            worksheet.unmerge_cells(range_string)
            shifted.append(
                f"{get_column_letter(min_col)}{min_row + amount}:"
                f"{get_column_letter(max_col)}{max_row + amount}"
            )
    return shifted


def copy_row_style(worksheet: Worksheet, source_row: int, target_row: int) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for col in range(1, worksheet.max_column + 1):
        source = worksheet.cell(row=source_row, column=col)
        target = worksheet.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def merge_property_row(worksheet: Worksheet, row: int) -> None:
    for start_col, end_col in (("A", "C"), ("I", "J")):
        range_string = f"{start_col}{row}:{end_col}{row}"
        if range_string not in [str(merged) for merged in worksheet.merged_cells.ranges]:
            worksheet.merge_cells(range_string)


def apply_property_rows(worksheet: Worksheet, data: dict[str, Any]) -> None:
    properties = normalize_property_items(data.get("imoveis"))
    if len(properties) == 1:
        first = properties[0]
        for key in (
            "area_total_ha",
            "area_pastagens_ha",
            "area_cultivo_ha",
            "area_financiada_bb_ha",
            "area_financiada_outros_ha",
            "atividade_principal",
            "principais_culturas",
        ):
            if first.get(key) in (None, "") and data.get(key) not in (None, ""):
                first[key] = data.get(key)
    if not properties:
        properties = [
            {
                "nome": data.get("imovel_nome"),
                "area_total_ha": data.get("area_total_ha"),
                "area_pastagens_ha": data.get("area_pastagens_ha"),
                "area_cultivo_ha": data.get("area_cultivo_ha"),
                "area_financiada_bb_ha": data.get("area_financiada_bb_ha"),
                "area_financiada_outros_ha": data.get("area_financiada_outros_ha"),
                "atividade_principal": data.get("atividade_principal"),
                "principais_culturas": data.get("principais_culturas"),
            }
        ]

    clear_property_rows(worksheet, max(PROPERTY_BASE_CAPACITY, len(properties)))
    for index, item in enumerate(properties):
        row = PROPERTY_START_ROW + index
        set_cell(worksheet, f"A{row}", upper_field(item.get("nome")))
        set_cell(worksheet, f"D{row}", item.get("area_total_ha"))
        set_cell(worksheet, f"E{row}", item.get("area_pastagens_ha"))
        set_cell(worksheet, f"F{row}", item.get("area_cultivo_ha"))
        set_cell(worksheet, f"G{row}", item.get("area_financiada_bb_ha"))
        set_cell(worksheet, f"H{row}", item.get("area_financiada_outros_ha"))
        set_cell(worksheet, f"I{row}", upper_field(item.get("atividade_principal")))
        set_cell(worksheet, f"K{row}", upper_field(item.get("principais_culturas")))


def clear_property_rows(worksheet: Worksheet, row_count: int) -> None:
    for row in range(PROPERTY_START_ROW, PROPERTY_START_ROW + row_count):
        for col in ("A", "D", "E", "F", "G", "H", "I", "K"):
            set_cell(worksheet, f"{col}{row}", None)


def shifted_coordinate(coordinate: str, row_offset: int, start_row: int = PROPERTY_SHIFT_START_ROW) -> str:
    if not row_offset:
        return coordinate
    match = re.fullmatch(r"([A-Z]+)(\d+)", coordinate)
    if not match:
        return coordinate
    col, row_text = match.groups()
    row = int(row_text)
    if row >= start_row:
        row += row_offset
    return f"{col}{row}"


def coordinate_row(coordinate: str) -> int | None:
    match = re.fullmatch(r"[A-Z]+(\d+)", coordinate)
    return int(match.group(1)) if match else None


def offset_for_coordinate(coordinate: str, row_offset: int, below_benfeitoria_offset: int) -> int:
    row = coordinate_row(coordinate)
    if row is not None and row >= BENFEITORIA_INSERT_AT_ROW:
        return below_benfeitoria_offset
    return row_offset


def shifted_row(row: int, row_offset: int, start_row: int = PROPERTY_SHIFT_START_ROW) -> int:
    return row + row_offset if row >= start_row else row


def clear_variable_model_values(worksheet: Worksheet, row_offset: int = 0) -> None:
    for cell in ["B4", "A207", *FIELD_TO_CELL.values(), *(cell for cell, _ in DATE_FIELDS.values())]:
        set_cell(worksheet, shifted_coordinate(cell, row_offset), None)

    for cell in ["A27", "A30", "A32", "A33", "F27", "G27", "H27", "I27", "F30", "G30", "H30", "I30", "F32", "G32", "H32", "I32"]:
        set_cell(worksheet, shifted_coordinate(cell, row_offset), None)

    for row in range(34, 41):
        for col in range(1, 12):
            set_cell(worksheet, f"{get_column_letter(col)}{shifted_row(row, row_offset)}", None)

    for row in range(46, EQUIPMENT_START_ROW):
        for col in range(1, 14):
            set_cell(worksheet, f"{get_column_letter(col)}{shifted_row(row, row_offset)}", None)

    clear_equipment_rows(worksheet, row_offset)
    clear_yes_no_area(worksheet, INSUMO_ROWS.values(), "E", "F", "G", row_offset)
    clear_yes_no_area(worksheet, PERSPECTIVA_ROWS.values(), "F", "G", "H", row_offset)
    clear_legacy_option_marks(worksheet, row_offset)


def apply_benfeitoria_blocks(worksheet: Worksheet, data: dict[str, Any], blocks: list[str], property_offset: int = 0) -> None:
    block_count = max(len(BENFEITORIA_BASE_BLOCKS), len(blocks))
    clear_benfeitoria_blocks(worksheet, block_count, property_offset)

    conservacao = normalize_key(str(data.get("benfeitorias_conservacao", "")))
    observacoes = data.get("benfeitorias_observacoes")
    for index in range(block_count):
        start_row, end_row = benfeitoria_block_rows(index, property_offset)
        format_benfeitoria_block(worksheet, start_row, end_row)
        has_text = index < len(blocks)
        if has_text:
            set_cell(worksheet, f"A{start_row}", blocks[index])
            style_benfeitoria_text_cell(worksheet, f"A{start_row}")
        if has_text and conservacao:
            set_mark(worksheet, f"F{start_row}", conservacao.startswith("bom"))
            set_mark(worksheet, f"G{start_row}", conservacao.startswith("regular"))
            set_mark(worksheet, f"H{start_row}", conservacao.startswith("ruim"))
        if has_text and observacoes not in (None, ""):
            set_cell(worksheet, f"I{start_row}", observacoes)
            style_benfeitoria_text_cell(worksheet, f"I{start_row}")
        adjust_benfeitoria_row_heights(worksheet, start_row, end_row, blocks[index] if index < len(blocks) else "")


def clear_benfeitoria_blocks(worksheet: Worksheet, block_count: int, property_offset: int = 0) -> None:
    for index in range(block_count):
        start_row, end_row = benfeitoria_block_rows(index, property_offset)
        unmerge_intersecting_range(worksheet, start_row, 1, end_row, 11)
        for row in range(start_row, end_row + 1):
            for col in range(1, 12):
                worksheet.cell(row=row, column=col).value = None


def benfeitoria_block_rows(index: int, property_offset: int = 0) -> tuple[int, int]:
    if index < len(BENFEITORIA_BASE_BLOCKS):
        start, end = BENFEITORIA_BASE_BLOCKS[index]
        return start + property_offset, end + property_offset

    extra_index = index - len(BENFEITORIA_BASE_BLOCKS)
    start = BENFEITORIA_INSERT_AT_ROW + property_offset + extra_index * BENFEITORIA_EXTRA_BLOCK_HEIGHT
    return start, start + BENFEITORIA_EXTRA_BLOCK_HEIGHT - 1


def format_benfeitoria_block(worksheet: Worksheet, start_row: int, end_row: int) -> None:
    unmerge_intersecting_range(worksheet, start_row, 1, end_row, 11)
    for range_string in (
        f"A{start_row}:E{end_row}",
        f"F{start_row}:F{end_row}",
        f"G{start_row}:G{end_row}",
        f"H{start_row}:H{end_row}",
        f"I{start_row}:K{end_row}",
    ):
        worksheet.merge_cells(range_string)
    apply_simple_range_border(worksheet, start_row, 1, end_row, 11)


def unmerge_intersecting_range(worksheet: Worksheet, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    for merged_range in list(worksheet.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if max_row < start_row or min_row > end_row or max_col < start_col or min_col > end_col:
            continue
        worksheet.unmerge_cells(str(merged_range))


def apply_simple_range_border(worksheet: Worksheet, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            current = copy(cell.border)
            cell.border = Border(
                left=BENFEITORIA_SIDE if col == start_col else current.left,
                right=BENFEITORIA_SIDE if col == end_col else current.right,
                top=BENFEITORIA_SIDE if row == start_row else current.top,
                bottom=BENFEITORIA_SIDE if row == end_row else current.bottom,
            )


def style_benfeitoria_text_cell(worksheet: Worksheet, coordinate: str) -> None:
    target = top_left_for_merged_cell(worksheet, coordinate)
    cell = worksheet[target]
    cell.alignment = copy(cell.alignment)
    cell.alignment = Alignment(
        horizontal=cell.alignment.horizontal,
        vertical="top",
        wrap_text=True,
    )


def adjust_benfeitoria_row_heights(worksheet: Worksheet, start_row: int, end_row: int, text: str) -> None:
    line_count = max(2, (len(text) // 95) + text.count("\n") + 1)
    total_height = max(42, min(150, line_count * 15 + 12))
    row_height = total_height / (end_row - start_row + 1)
    for row in range(start_row, end_row + 1):
        worksheet.row_dimensions[row].height = max(18, row_height)


def apply_fields(worksheet: Worksheet, data: dict[str, Any], row_offset: int = 0, below_benfeitoria_offset: int | None = None) -> None:
    below_benfeitoria_offset = row_offset if below_benfeitoria_offset is None else below_benfeitoria_offset
    client_block = build_client_block(data)
    if client_block:
        set_cell(worksheet, "B4", client_block)

    if not data.get("finalidade_vistoria"):
        data["finalidade_vistoria"] = "VERIFICAÇÃO IN-LOCO DE REAIS CONDIÇÕES DE PRODUTIVIDADE DO CLIENTE EM QUESTÃO;"

    property_fields = {
        "imovel_nome",
        "area_total_ha",
        "area_pastagens_ha",
        "area_cultivo_ha",
        "area_financiada_bb_ha",
        "area_financiada_outros_ha",
        "atividade_principal",
        "principais_culturas",
    }
    for field, cell in FIELD_TO_CELL.items():
        if field in property_fields:
            continue
        if field in {"benfeitorias_descricao", "benfeitorias_observacoes"}:
            continue
        value = data.get(field)
        if value not in (None, ""):
            offset = offset_for_coordinate(cell, row_offset, below_benfeitoria_offset)
            set_cell(worksheet, shifted_coordinate(cell, offset), value)

    for field, (cell, pattern) in DATE_FIELDS.items():
        value = data.get(field)
        if value not in (None, ""):
            offset = offset_for_coordinate(cell, row_offset, below_benfeitoria_offset)
            set_cell(worksheet, shifted_coordinate(cell, offset), pattern.format(value))
    if data.get("data_visita") not in (None, ""):
        set_cell(worksheet, shifted_coordinate("A207", below_benfeitoria_offset), f"DATA DA VISITA: {data['data_visita']}")

    if data.get("raw_text") and not data.get("outros_comentarios"):
        set_cell(worksheet, shifted_coordinate("B181", below_benfeitoria_offset), data["raw_text"])


def build_client_block(data: dict[str, Any]) -> str | None:
    if data.get("resumo_cliente"):
        return upper_field(str(data["resumo_cliente"]))

    if data.get("cliente"):
        return upper_field(str(data["cliente"]).strip()) or None

    return None


def upper_field(value: Any) -> Any:
    """Caixa alta para os campos da tabela de discriminacao (cliente, imovel,
    atividade, culturas), seguindo o padrao dos modelos aprovados. Nao se aplica
    aos textos de benfeitorias/comentarios/conclusao, que ficam em caixa normal."""
    return value.upper() if isinstance(value, str) else value


def build_property_line(data: dict[str, Any]) -> str | None:
    property_names = property_names_from_value(data.get("imoveis"))
    name = format_property_names(property_names) if property_names else data.get("imovel_nome")
    if not name:
        return None

    area_parts = []
    if data.get("area_alqueires"):
        area_parts.append(f"{data['area_alqueires']} alqueires")
    if data.get("area_total_ha"):
        area_parts.append(f"{data['area_total_ha']} ha")

    if area_parts:
        return f"{name} - {' / '.join(area_parts)}"
    return str(name)


def enrich_location_summary(data: dict[str, Any]) -> None:
    city = data.get("cidade_uf")
    location = data.get("localizacao_1")
    property_name = data.get("imovel_resumo") or data.get("imovel_nome")
    if not city or not property_name:
        return
    if location not in (None, "", city):
        return

    data["localizacao_1"] = f"{city} - {property_name}"


def apply_equipment(worksheet: Worksheet, equipment: Any, row_offset: int = 0) -> None:
    clear_equipment_rows(worksheet, row_offset)
    if not isinstance(equipment, list):
        return

    for index, item in enumerate(equipment[: EQUIPMENT_END_ROW - EQUIPMENT_START_ROW + 1]):
        if not isinstance(item, dict):
            continue
        row = shifted_row(EQUIPMENT_START_ROW + index, row_offset)
        normalized_item = normalize_nested_dict(item)
        for field, col in EQUIPMENT_COLUMNS.items():
            value = normalized_item.get(field)
            if value not in (None, ""):
                set_cell(worksheet, f"{col}{row}", value)


def infer_insumos(data: dict[str, Any]) -> dict[str, bool]:
    """Marca insumos SOMENTE quando ha evidencia explicita no relatorio.

    Nunca inventa: agua so com mencao a bebedouro/poco/corrego/tanque/represa/
    nascente/acude/piscicultura; energia so com placa solar/fotovoltaica/rede
    eletrica/gerador; pastagens so com mencao a pasto/piquete/capim/braquiaria.
    """
    blob = _gather_evidence_text(data)
    if not blob:
        return {}

    evidence = {
        "agua": (
            "bebedouro",
            "poco",
            "artesiano",
            "corrego",
            "represa",
            "tanque",
            "nascente",
            "acude",
            "piscicultura",
            "lamina_d",
            "reservatorio",
        ),
        "energia_eletrica": (
            "placa_solar",
            "placas_solar",
            "fotovoltaic",
            "energia_eletrica",
            "rede_eletrica",
            "trifasic",
            "gerador",
            "painel_solar",
        ),
        "pastagens": (
            "pastagem",
            "pasto",
            "piquete",
            "braquiar",
            "brachiar",
            "andropogon",
            "quicuia",
            "mombaca",
            "capim",
        ),
        "mao_de_obra": (
            "funcionario",
            "funcionarios",
            "caseiro",
            "trabalhador",
            "trabalhadores",
            "alojamento",
            "colaborador",
            "vaqueiro",
            "peao",
            "peoes",
        ),
        "estrutura_armazenagem": (
            "galpao",
            "silo",
            "armazem",
            "barracao",
            "trincheira",
            "deposito",
            "paiol",
            "tulha",
        ),
        "estrutura_transporte": (
            "caminhao",
            "caminhoes",
            "caminhonete",
            "carreta",
            "frota",
        ),
    }
    marks: dict[str, bool] = {}
    for key, terms in evidence.items():
        if any(term in blob for term in terms):
            marks[key] = True
    return marks


def _gather_evidence_text(data: dict[str, Any]) -> str:
    parts: list[str] = []
    for field in (
        "benfeitorias_descricao",
        "benfeitorias_observacoes",
        "outros_comentarios",
        "investimentos_comentarios",
        "insumos_comentarios",
        "conclusao",
        "principais_culturas",
        "atividade_principal",
        "raw_text",
    ):
        value = data.get(field)
        if value:
            parts.append(str(value))
    for item in (data.get("imoveis") or []):
        if isinstance(item, dict):
            parts.extend(str(item.get(key, "")) for key in ("atividade_principal", "principais_culturas", "nome"))
    return normalize_key(_drop_negated_clauses(" ".join(parts)))


def _drop_negated_clauses(text: str) -> str:
    """Remove trechos negados para a rede de seguranca nao marcar por engano.

    Ex.: 'nao ha energia eletrica' ou 'sem poco artesiano' nao devem virar
    evidencia de energia/agua.
    """
    negation = re.compile(r"\b(?:n[aã]o|sem|inexist\w*|aus[eê]nc\w*|destitu\w*)\b", re.IGNORECASE)
    kept: list[str] = []
    for clause in re.split(r"[.;,\n]", text):
        if not negation.search(clause):
            kept.append(clause)
    return " ".join(kept)


def apply_insumos(worksheet: Worksheet, insumos: Any, row_offset: int = 0) -> None:
    clear_yes_no_area(worksheet, INSUMO_ROWS.values(), "E", "F", "G", row_offset)
    if not isinstance(insumos, dict):
        return

    for key, row in INSUMO_ROWS.items():
        item = insumos.get(key)
        if item is None:
            continue
        apply_yes_no_observation(worksheet, shifted_row(row, row_offset), "E", "F", "G", item)


def apply_perspectivas(worksheet: Worksheet, perspectivas: Any, row_offset: int = 0) -> None:
    clear_yes_no_area(worksheet, PERSPECTIVA_ROWS.values(), "F", "G", "H", row_offset)
    if not isinstance(perspectivas, dict):
        return

    normalized = normalize_nested_dict(perspectivas)
    for key, row in PERSPECTIVA_ROWS.items():
        item = normalized.get(key)
        if item is None:
            continue
        apply_yes_no_observation(worksheet, shifted_row(row, row_offset), "F", "G", "H", item)


def apply_yes_no_observation(
    worksheet: Worksheet,
    row: int,
    yes_col: str,
    no_col: str,
    observation_col: str,
    item: Any,
) -> None:
    value = item
    observation = None

    if isinstance(item, dict):
        normalized_item = normalize_nested_dict(item)
        value = normalized_item.get("sim", normalized_item.get("valor", normalized_item.get("status")))
        observation = normalized_item.get("observacao", normalized_item.get("observacoes"))

    truthy = parse_bool(value)
    set_mark(worksheet, f"{yes_col}{row}", truthy is True)
    set_mark(worksheet, f"{no_col}{row}", truthy is False)
    if observation not in (None, ""):
        set_cell(worksheet, f"{observation_col}{row}", observation)


def parse_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    normalized = normalize_key(str(value))
    if normalized in {"sim", "s", "true", "x", "1", "yes"}:
        return True
    if normalized in {"nao", "n", "false", "0", "no"}:
        return False
    return None


def set_cell(worksheet: Worksheet, coordinate: str, value: Any) -> None:
    target = top_left_for_merged_cell(worksheet, coordinate)
    cell = worksheet[target]
    cell.value = value
    if value not in (None, ""):
        cell.alignment = copy(cell.alignment)
        cell.alignment = Alignment(
            horizontal="left",
            vertical=cell.alignment.vertical or "center",
            wrap_text=cell.alignment.wrap_text,
        )


def set_mark(worksheet: Worksheet, coordinate: str, enabled: bool) -> None:
    target = top_left_for_merged_cell(worksheet, coordinate)
    cell = worksheet[target]
    cell.value = "X" if enabled else None
    cell.alignment = copy(cell.alignment)
    cell.alignment = Alignment(
        horizontal="center",
        vertical=cell.alignment.vertical or "center",
        wrap_text=cell.alignment.wrap_text,
    )
    cell.font = copy(cell.font)
    cell.font = Font(
        name=cell.font.name,
        size=cell.font.sz,
        bold=True,
        italic=cell.font.italic,
        color=cell.font.color,
        underline=cell.font.underline,
    )


def top_left_for_merged_cell(worksheet: Worksheet, coordinate: str) -> str:
    for merged_range in worksheet.merged_cells.ranges:
        if coordinate in merged_range:
            return merged_range.start_cell.coordinate
    return coordinate


def clear_equipment_rows(worksheet: Worksheet, row_offset: int = 0) -> None:
    rows = [40, *range(EQUIPMENT_START_ROW, EQUIPMENT_END_ROW + 2)]
    rows_to_clear = {row for row in rows}
    rows_to_clear.update(shifted_row(row, row_offset) for row in rows)
    for shifted in rows_to_clear:
        for col in EQUIPMENT_COLUMNS.values():
            set_cell(worksheet, f"{col}{shifted}", None)


def clear_yes_no_area(worksheet: Worksheet, rows: Any, yes_col: str, no_col: str, observation_col: str, row_offset: int = 0) -> None:
    for row in rows:
        shifted = shifted_row(row, row_offset)
        for col in (yes_col, no_col, observation_col):
            set_cell(worksheet, f"{col}{shifted}", None)


def clear_legacy_option_marks(worksheet: Worksheet, row_offset: int = 0) -> None:
    option_cells = [
        "A150",
        "D150",
        "H150",
        "A151",
        "D151",
        "H151",
        *[f"B{row}" for row in range(154, 159)],
        *[f"B{row}" for row in range(161, 164)],
        *[f"B{row}" for row in range(166, 169)],
        "B171",
        "B172",
        "B175",
        "B176",
    ]
    for coordinate in option_cells:
        target = top_left_for_merged_cell(worksheet, shifted_coordinate(coordinate, row_offset))
        value = worksheet[target].value
        if isinstance(value, str):
            worksheet[target].value = re.sub(r"\(\s*[xX]\s*\)", "(   )", value)

    for coordinate in ["B172", "B176"]:
        target = top_left_for_merged_cell(worksheet, shifted_coordinate(coordinate, row_offset))
        value = worksheet[target].value
        if isinstance(value, str) and ":" in value:
            prefix = value.split(":", 1)[0].rstrip()
            worksheet[target].value = f"{prefix}: "


def polish_written_ranges(worksheet: Worksheet, row_offset: int = 0, below_benfeitoria_offset: int | None = None) -> None:
    below_benfeitoria_offset = row_offset if below_benfeitoria_offset is None else below_benfeitoria_offset
    wrap_cells = [
        "B4",
        "C8",
        "A12",
        "A18",
        "I18",
        "K18",
        "A27",
        "I27",
        "D96",
        "D108",
        "D132",
        "B147",
        "B184",
        "B193",
    ]
    for coordinate in wrap_cells:
        offset = offset_for_coordinate(coordinate, row_offset, below_benfeitoria_offset)
        target = top_left_for_merged_cell(worksheet, shifted_coordinate(coordinate, offset))
        cell = worksheet[target]
        cell.alignment = copy(cell.alignment)
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal,
            vertical=cell.alignment.vertical or "center",
            wrap_text=True,
        )

    for row in range(PROPERTY_START_ROW, PROPERTY_TEMPLATE_END_ROW + row_offset + 1):
        for col in ("D", "E", "F", "G", "H"):
            target = top_left_for_merged_cell(worksheet, f"{col}{row}")
            worksheet[target].number_format = "0.00"


def adjust_dynamic_row_heights(worksheet: Worksheet) -> None:
    header_value = worksheet[top_left_for_merged_cell(worksheet, "B4")].value
    if isinstance(header_value, str) and header_value.strip():
        line_count = header_value.count("\n") + 1
        worksheet.row_dimensions[4].height = max(96, min(210, line_count * 18 + 18))


def apply_photos(worksheet: Worksheet, photo_paths: list[str | Path], image_output_dir: Path, row_offset: int = 0) -> None:
    valid_photos = unique_existing_photo_paths(photo_paths)
    remove_old_report_photos(worksheet)
    clear_photo_slots(worksheet, row_offset)
    if not valid_photos:
        return

    image_output_dir.mkdir(parents=True, exist_ok=True)
    for old_photo in image_output_dir.glob("foto-*.jpg"):
        old_photo.unlink(missing_ok=True)

    for index, photo_path in enumerate(valid_photos[: len(PHOTO_ANCHORS)]):
        from_row, from_col, to_row, to_col = PHOTO_ANCHORS[index]
        from_row = shifted_row(from_row + 1, row_offset) - 1
        to_row = shifted_row(to_row + 1, row_offset) - 1
        format_photo_slot(worksheet, index + 1, from_row, from_col, to_row, to_col)

        prepared_path = prepare_photo(photo_path, image_output_dir, index + 1)
        image = ExcelImage(str(prepared_path))
        image.width, image.height = PHOTO_MAX_SIZE
        image.anchor = f"{get_column_letter(from_col + 1)}{from_row + 2}"
        worksheet.add_image(image)


def unique_existing_photo_paths(photo_paths: list[str | Path]) -> list[Path]:
    unique: list[Path] = []
    seen: set[str] = set()
    for path_value in photo_paths:
        path = Path(path_value)
        if not path.exists() or not path.is_file():
            continue
        signature = photo_signature(path)
        if signature in seen:
            continue
        seen.add(signature)
        unique.append(path)
    return unique


def photo_signature(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clear_photo_slots(worksheet: Worksheet, row_offset: int = 0) -> None:
    for from_row, from_col, to_row, to_col in PHOTO_ANCHORS:
        from_row = shifted_row(from_row + 1, row_offset) - 1
        to_row = shifted_row(to_row + 1, row_offset) - 1
        for row in range(from_row + 1, to_row + 3):
            for col in range(from_col + 1, to_col + 2):
                cell = worksheet.cell(row=row, column=col)
                cell.value = None
                cell.border = Border()
                cell.fill = PatternFill(fill_type=None)


def format_photo_slot(worksheet: Worksheet, number: int, from_row: int, from_col: int, to_row: int, to_col: int) -> None:
    start_row = from_row + 1
    end_row = to_row + 1
    start_col = from_col + 1
    end_col = to_col + 1

    caption_cell = worksheet.cell(row=start_row, column=start_col)
    caption_cell.value = f"Foto {number:02d}"
    caption_cell.font = copy(caption_cell.font)
    caption_cell.font = Font(
        name=caption_cell.font.name,
        size=caption_cell.font.sz or 10,
        bold=True,
        italic=caption_cell.font.italic,
        color="173B2C",
        underline=caption_cell.font.underline,
    )
    caption_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[start_row].height = max(18, worksheet.row_dimensions[start_row].height or 0)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            worksheet.cell(row=row, column=col).border = Border()


def remove_old_report_photos(worksheet: Worksheet) -> None:
    kept_images = []
    for image in getattr(worksheet, "_images", []):
        row = image_anchor_row(image)
        if row is None or row < 200:
            kept_images.append(image)
    worksheet._images = kept_images


def image_anchor_row(image: ExcelImage) -> int | None:
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    return int(marker.row) + 1


def prepare_photo(photo_path: Path, output_dir: Path, index: int) -> Path:
    output_path = output_dir / f"foto-{index:02d}.jpg"
    with Image.open(photo_path) as original:
        image = ImageOps.exif_transpose(original).convert("RGB")
        inner_size = (
            max(1, PHOTO_MAX_SIZE[0] - PHOTO_BORDER_PX * 2),
            max(1, PHOTO_MAX_SIZE[1] - PHOTO_BORDER_PX * 2),
        )
        fitted = image.resize(inner_size, Image.Resampling.LANCZOS)
        framed = Image.new("RGB", PHOTO_MAX_SIZE, PHOTO_BORDER_COLOR)
        framed.paste(fitted, (PHOTO_BORDER_PX, PHOTO_BORDER_PX))
        framed.save(output_path, quality=88, optimize=True)
    return output_path


def collect_photos(photo_dir: str | Path | None) -> list[Path]:
    if not photo_dir:
        return []
    root = Path(photo_dir)
    if not root.exists():
        return []
    allowed = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}
    return sorted(path for path in root.iterdir() if path.suffix.lower() in allowed)


def load_data_argument(data_arg: str | None, data_file: str | None) -> str:
    if data_file:
        return Path(data_file).read_text(encoding="utf-8")
    return data_arg or ""


def copy_template_to_workspace(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera relatorio agronomico em Excel.")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="Caminho do modelo .xlsx")
    parser.add_argument("--data", default=None, help="JSON ou texto com campos no formato campo: valor")
    parser.add_argument("--data-file", default=None, help="Arquivo .json/.txt com os dados")
    parser.add_argument("--photos", default=None, help="Pasta com fotos")
    parser.add_argument("--out", default=None, help="Caminho do .xlsx gerado")
    args = parser.parse_args()

    data_text = load_data_argument(args.data, args.data_file)
    photos = collect_photos(args.photos)
    output = generate_report(data_text, photos, args.out, args.template)
    print(output)


if __name__ == "__main__":
    main()
