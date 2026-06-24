from __future__ import annotations

import json
from pathlib import Path
import sys

from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from relatorio_app.pattern_library import build_writer_prompt, classify_case_tags, load_pattern_examples, select_pattern_examples
from relatorio_app.report_engine import generate_report, parse_decimal_pt, parse_report_data
from relatorio_app.technical_writer import generate_technical_report
from server import parse_review_data

OUTPUT_DIR = ROOT / "outputs" / "validation"


CASES = [
    {
        "name": "wesley",
        "source": ROOT / "sample_gemini_text.txt",
        "output": OUTPUT_DIR / "wesley.xlsx",
        "fields": {
            "cliente": "Wesley Paulo",
            "cidade_uf": "Cromínia-GO",
            "imovel_nome": "Sítio Pedro Rosa (Fazenda Santa Bárbara)",
            "area_total_ha": 53.24,
            "area_pastagens_ha": 37.27,
            "area_cultivo_ha": 15.97,
            "atividade_principal": "Pecuária mista (Leite e Corte)",
            "principais_culturas": "Pastagens forrageiras destinadas ao pastejo do rebanho leiteiro e de corte",
        },
        "cells": {
            "A18": "Sítio Pedro Rosa (Fazenda Santa Bárbara)",
            "D18": 53.24,
            "E18": 37.27,
            "F18": 15.97,
            "F27": "X",
            "A51": "Não informado",
            "G40": None,
            "F30": None,
            "I30": None,
        },
    },
    {
        "name": "mandioca",
        "source": ROOT / "sample_mandioca_text.txt",
        "output": OUTPUT_DIR / "mandioca.xlsx",
        "fields": {
            "cliente": "Lúcio Mauro",
            "imovel_nome": "Fazenda Água Branca",
            "area_total_ha": 29.04,
            "area_pastagens_ha": 0.0,
            "area_cultivo_ha": 29.04,
            "atividade_principal": "Cultivo comercial de Mandioca (Manihot esculenta)",
            "principais_culturas": "Mandioca",
        },
        "cells": {
            "A18": "Fazenda Água Branca",
            "D18": 29.04,
            "E18": 0,
            "F18": 29.04,
            "F27": "X",
            "A51": "Não informado",
            "G40": None,
            "F30": None,
            "I30": None,
        },
    },
]


MULTIPLE_PROPERTY_TEXT = """
1. DISCRIMINACAO
Nomes das propriedades: Fazenda Boa Vista; Fazenda Santa Luzia
Area Total (ha): 120,50 ha
Area de Pastagens (ha): 80,00 ha
Area de Cultivo (ha): 40,50 ha
Atividade principal desenvolvida: Pecuaria de corte
Principais culturas: Pastagens
2. TIPO (Benfeitorias e Infraestrutura)
O produtor possui as fazendas Fazenda Boa Vista e Fazenda Santa Luzia localizadas no mesmo municipio.
OUTROS COMENTARIOS
As areas sao conduzidas de forma integrada.
CONCLUSAO
Atividade regular.
"""


SANDRO_STYLE_TEXT = """
1. DISCRIMINACAO
Nome da propriedade: Complexo Agropecuario Teste (Fazendas: Boa Sorte, Sao Judas, Santa Camila, Sao Jose dos Anjicos e Santa Carolina)

Dados de Area e Exploracao por Propriedade:

Fazenda Boa Sorte
Area Total (ha): 3.388,00 ha
Area de Pastagens (ha): 2.371,60 ha
Area de Cultivo (ha): 1.016,40 ha
Atividade principal desenvolvida: Pecuaria de corte (Recria)
Principais culturas: Pastagens de Andropogon, Quicuia e Braquiarao

Fazenda Sao Judas
Area Total (ha): 2.429,68 ha
Area de Pastagens (ha): 1.700,78 ha
Area de Cultivo (ha): 728,90 ha
Atividade principal desenvolvida: Pecuaria de corte (Cria)
Principais culturas: Pastagens de Andropogon

Fazenda Santa Camila
Area Total (ha): 1.863,40 ha
Area de Pastagens (ha): 1.304,38 ha
Area de Cultivo (ha): 559,02 ha
Atividade principal desenvolvida: Pecuaria de corte (Cria)
Principais culturas: Pastagens de Quicuia

Fazenda Sao Jose dos Anjicos
Area Total (ha): 731,00 ha
Area de Pastagens (ha): 511,70 ha
Area de Cultivo (ha): 219,30 ha
Atividade principal desenvolvida: Pecuaria de corte (Recria estruturada)
Principais culturas: Pastagens de Andropogon

Fazenda Santa Carolina
Area Total (ha): 821,00 ha
Area de Pastagens (ha): 574,70 ha
Area de Cultivo (ha): 246,30 ha
Atividade principal desenvolvida: Pecuaria de corte (Terminacao)
Principais culturas: Pastagens de Andropogon

2. TIPO (Benfeitorias e Infraestrutura)
O complexo pecuario apresenta infraestrutura integrada e alto padrao tecnico.

INVESTIMENTOS EM ANDAMENTO (Comentarios)
Projeto com cinco unidades produtivas e manejo integrado.

OUTROS COMENTARIOS
Operacao conduzida em grande escala com controles mensais.

CONCLUSAO
Atividade regular e tecnicamente recomendada.

FRASES DIRETAS
COMPLEXO PECUARIO COM CINCO UNIDADES PRODUTIVAS INTEGRADAS.
"""


COLLAPSED_PROPERTY_TEXT = (
    "1. DISCRIMINACAONome da propriedade: Fazenda Beira Rio e Fazenda Val da Onca"
    "Tipo de exploracao: Propria"
    "Dados de Area e Exploracao por Propriedade:"
    "Fazenda Beira Rio"
    "Area Total (ha): 580,80 ha"
    "Area de Pastagens (ha): 406,56 ha"
    "Area de Cultivo (ha): 174,24 ha"
    "Atividade principal desenvolvida: Pecuaria mista"
    "Principais culturas: Pastagens forrageiras"
    "Fazenda Val da Onca"
    "Area Total (ha): 580,80 ha"
    "Area de Pastagens (ha): 406,56 ha"
    "Area de Cultivo (ha): 174,24 ha"
    "Atividade principal desenvolvida: Pecuaria mista"
    "Principais culturas: Pastagens forrageiras"
    "2. TIPO (Benfeitorias e Infraestrutura)"
    "O projeto agropecuario fica em duas unidades produtivas."
    "INVESTIMENTOS EM ANDAMENTO (Comentarios)"
    "Investimentos integrados."
    "OUTROS COMENTARIOS"
    "Comentarios tecnicos."
    "CONCLUSAO"
    "Atividade regular."
    "FRASES DIRETAS"
    "OPERACAO INTEGRADA EM DUAS FAZENDAS."
)


COLLAPSED_GROUP_TEXT = (
    "1. DISCRIMINACAONome da propriedade: Grupo Pedra (Unidades: Buriti e Aroeira)"
    "Dados de Area e Exploracao por Propriedade:"
    "Grupo Pedra - Unidade Buriti (Ituverava-SP)"
    "Area Total (ha): 1.432,35 ha"
    "Area de Pastagens (ha): 0,00 ha"
    "Area de Cultivo (ha): 1.432,35 ha"
    "Atividade principal desenvolvida: Cultivo comercial de Cana-de-Acucar"
    "Principais culturas: Cana-de-Acucar"
    "Grupo Pedra - Unidade Aroeira (Tupaciguara-MG)"
    "Area Total (ha): 1.800,00 ha"
    "Area de Pastagens (ha): 0,00 ha"
    "Area de Cultivo (ha): 1.800,00 ha"
    "Atividade principal desenvolvida: Cultivo comercial de Cana-de-Acucar e graos"
    "Principais culturas: Cana-de-Acucar, Soja e Amendoim"
    "2. TIPO (Benfeitorias e Infraestrutura)"
    "O complexo agricola apresenta estrutura operacional robusta."
)


DETAIL_PROPERTY_TEXT = (
    "1. DISCRIMINACAONome da propriedade: Fazenda Sao Jose e Fazenda Santo Antonio"
    "Dados de Area e Exploracao:"
    "Area Total (ha): 329,32 ha"
    "Area de Pastagens (ha): 230,52 ha"
    "Area de Cultivo (ha): 98,80 ha"
    "Atividade principal desenvolvida: Pecuaria de corte"
    "Principais culturas: Pastagens"
    "Detalhamento por Unidade (ha):"
    "Fazenda Sao Jose (Jaupaci-GO - 50,04 alqueires): 242,19 ha brutos / 169,53 ha de pastagens liquidas"
    "Fazenda Santo Antonio (Palmeiras-GO - 18,00 alqueires): 87,12 ha brutos / 60,99 ha de pastagens liquidas"
    "2. TIPO (Benfeitorias e Infraestrutura)"
    "As unidades possuem infraestrutura produtiva."
)


BENFEITORIA_SPLIT_TEXT = (
    "1. DISCRIMINACAONome da propriedade: Complexo Agropecuario Sandro Mabel "
    "(Fazendas: Boa Sorte e Sao Judas)"
    "Dados de Area e Exploracao por Propriedade:"
    "Fazenda Boa Sorte"
    "Area Total (ha): 100,00 ha"
    "Area de Pastagens (ha): 70,00 ha"
    "Area de Cultivo (ha): 30,00 ha"
    "Atividade principal desenvolvida: Pecuaria de corte"
    "Principais culturas: Pastagens"
    "Fazenda Sao Judas"
    "Area Total (ha): 200,00 ha"
    "Area de Pastagens (ha): 140,00 ha"
    "Area de Cultivo (ha): 60,00 ha"
    "Atividade principal desenvolvida: Pecuaria de corte"
    "Principais culturas: Pastagens"
    "2. TIPO (Benfeitorias e Infraestrutura)"
    "O complexo pecuario gerido pelo produtor Sandro Mabel em Sao Miguel do Araguaia-GO "
    "possui infraestrutura geral integrada."
    "Na Fazenda Boa Sorte, ha 60 pastos com cochos cobertos e bebedouros automatizados."
    "Na Fazenda Sao Judas, o rebanho e manejado em pastagens de Andropogon."
    "INVESTIMENTOS EM ANDAMENTO (Comentarios)"
    "Sem investimentos relevantes."
)


def assert_equal(actual, expected, label: str) -> None:
    if isinstance(expected, float):
        if round(float(actual), 2) != round(expected, 2):
            raise AssertionError(f"{label}: esperado {expected!r}, veio {actual!r}")
        return
    if actual != expected:
        raise AssertionError(f"{label}: esperado {expected!r}, veio {actual!r}")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    for raw, expected in {"29,04 ha": 29.04, "29.04": 29.04, "1.234,56": 1234.56}.items():
        assert_equal(parse_decimal_pt(raw), expected, f"conversão numérica {raw}")

    multiple = parse_report_data(MULTIPLE_PROPERTY_TEXT)
    expected_properties = [{"nome": "Fazenda Boa Vista"}, {"nome": "Fazenda Santa Luzia"}]
    assert_equal(multiple.get("imovel_nome"), "Fazenda Boa Vista; Fazenda Santa Luzia", "multiplas propriedades campo imovel")
    assert_equal(multiple.get("imoveis"), expected_properties, "multiplas propriedades lista")

    multiple_output = generate_report(MULTIPLE_PROPERTY_TEXT, output_path=OUTPUT_DIR / "multiplas-propriedades.xlsx")
    multiple_worksheet = load_workbook(multiple_output).active
    assert_equal(multiple_worksheet["A18"].value, "Fazenda Boa Vista", "primeira propriedade celula A18")
    assert_equal(multiple_worksheet["A19"].value, "Fazenda Santa Luzia", "segunda propriedade celula A19")

    phrase = parse_report_data("O produtor explora as fazendas Fazenda Primavera e Fazenda Santa Clara localizadas em Morrinhos-GO.")
    assert_equal(phrase.get("imoveis"), [{"nome": "Fazenda Primavera"}, {"nome": "Fazenda Santa Clara"}], "multiplas propriedades em frase")

    from_json = parse_report_data({"imoveis": [{"nome": "Fazenda Modelo"}, {"nome": "Sitio Dois Irmaos"}]})
    assert_equal(from_json.get("imovel_nome"), "Fazenda Modelo; Sitio Dois Irmaos", "multiplas propriedades em JSON")

    sandro_style = parse_report_data(SANDRO_STYLE_TEXT)
    assert_equal(len(sandro_style.get("imoveis", [])), 5, "sandro quantidade de propriedades")
    assert_equal(sandro_style["imoveis"][0]["area_total_ha"], 3388.0, "sandro area primeira propriedade")
    assert_equal(sandro_style.get("area_total_ha"), 9233.08, "sandro area total consolidada")

    sandro_output = generate_report(SANDRO_STYLE_TEXT, output_path=OUTPUT_DIR / "sandro-multiplas-propriedades.xlsx")
    sandro_worksheet = load_workbook(sandro_output).active
    assert_equal(sandro_worksheet["A18"].value, "Fazenda Boa Sorte", "sandro propriedade linha 18")
    assert_equal(sandro_worksheet["A19"].value, "Fazenda Sao Judas", "sandro propriedade linha 19")
    assert_equal(sandro_worksheet["A20"].value, "Fazenda Santa Camila", "sandro propriedade linha 20")
    assert_equal(sandro_worksheet["A21"].value, "Fazenda Sao Jose dos Anjicos", "sandro propriedade linha 21")
    assert_equal(sandro_worksheet["A22"].value, "Fazenda Santa Carolina", "sandro propriedade linha 22")
    assert_equal(sandro_worksheet["D22"].value, 821.0, "sandro area quinta propriedade")
    assert_equal(sandro_worksheet["A24"].value, "3        BENFEITORIAS", "sandro secao benfeitorias deslocada")
    assert_equal(sandro_worksheet["A28"].value, "O complexo pecuario apresenta infraestrutura integrada e alto padrao tecnico.", "sandro texto benfeitorias deslocado")
    assert_equal(sandro_worksheet["D97"].value, "Projeto com cinco unidades produtivas e manejo integrado.", "sandro investimentos deslocado")
    assert_equal(sandro_worksheet["D109"].value, "COMPLEXO PECUARIO COM CINCO UNIDADES PRODUTIVAS INTEGRADAS.", "sandro frase direta deslocada")

    collapsed = parse_report_data(COLLAPSED_PROPERTY_TEXT)
    assert_equal(collapsed.get("imovel_nome"), "Fazenda Beira Rio; Fazenda Val da Onca", "texto colado nomes limpos")
    assert_equal(collapsed["imoveis"][0]["principais_culturas"], "Pastagens forrageiras", "texto colado culturas primeira propriedade")
    assert_equal(collapsed["imoveis"][1]["nome"], "Fazenda Val da Onca", "texto colado segunda propriedade limpa")
    assert_equal(collapsed.get("benfeitorias_descricao"), "O projeto agropecuario fica em duas unidades produtivas.", "texto colado benfeitorias separadas")

    collapsed_group = parse_report_data(COLLAPSED_GROUP_TEXT)
    assert_equal(collapsed_group.get("imovel_nome"), "Grupo Pedra - Unidade Buriti (Ituverava-SP); Grupo Pedra - Unidade Aroeira (Tupaciguara-MG)", "grupo colado nomes limpos")
    assert_equal(collapsed_group["imoveis"][1]["principais_culturas"], "Cana-de-Acucar, Soja e Amendoim", "grupo colado culturas segunda unidade")
    assert_equal(collapsed_group.get("area_total_ha"), 3232.35, "grupo colado area consolidada")

    detail = parse_report_data(DETAIL_PROPERTY_TEXT)
    assert_equal(detail.get("area_total_ha"), 329.32, "detalhamento preserva area consolidada")
    assert_equal(detail["imoveis"][0]["area_total_ha"], 242.19, "detalhamento primeira unidade")
    assert_equal(detail["imoveis"][1]["area_cultivo_ha"], 26.13, "detalhamento calcula cultivo segunda unidade")

    benfeitoria_split = parse_report_data(BENFEITORIA_SPLIT_TEXT)
    assert_equal(benfeitoria_split.get("cliente"), "Sandro Mabel", "nome do cliente extraido de paragrafo tecnico")
    benfeitoria_output = generate_report(BENFEITORIA_SPLIT_TEXT, output_path=OUTPUT_DIR / "benfeitorias-em-caixas.xlsx")
    benfeitoria_worksheet = load_workbook(benfeitoria_output).active
    assert_equal(benfeitoria_worksheet["B4"].value, "Sandro Mabel", "cliente escrito somente como nome")
    assert_equal(benfeitoria_worksheet["A27"].value.startswith("Na Fazenda Boa Sorte"), True, "benfeitorias primeira fazenda")
    assert_equal(benfeitoria_worksheet["A30"].value.startswith("Na Fazenda Sao Judas"), True, "benfeitorias segunda fazenda")
    assert_equal(
        "complexo pecuario" in (benfeitoria_worksheet["B184"].value or "").lower(),
        True,
        "visao geral das benfeitorias em outros comentarios",
    )

    manual_text = (
        "Cliente: Maria Souza\n"
        "Data da visita: 15/06/2026\n"
        "Maquinarios: Trator Massey Ferguson 4292; Grade aradora 16 discos\n"
    )
    manual_output = generate_report(manual_text, output_path=OUTPUT_DIR / "manual-data-maquinas.xlsx")
    manual_worksheet = load_workbook(manual_output).active
    assert_equal(manual_worksheet["J4"].value, "Data da Visita:15/06/2026", "data da visita preenchida")
    assert_equal(manual_worksheet["A195"].value, "Data: 15/06/2026", "data assinatura preenchida")
    assert_equal(manual_worksheet["A204"].value, "Data: 15/06/2026", "data administracao preenchida")
    assert_equal(manual_worksheet["A207"].value, "DATA DA VISITA: 15/06/2026", "data rodape preenchida")
    assert_equal(manual_worksheet["A51"].value, "Trator Massey Ferguson 4292", "maquinario primeira linha")
    assert_equal(manual_worksheet["A52"].value, "Grade aradora 16 discos", "maquinario segunda linha")
    assert_equal(manual_worksheet["A51"].alignment.horizontal, "left", "maquinario alinhado a esquerda")

    raw_arnaldo = (ROOT / "sample_raw_arnaldo.txt").read_text(encoding="utf-8")
    technical_result = generate_technical_report(raw_arnaldo)
    technical_parsed = parse_report_data(technical_result.report_text)
    pattern_tags = classify_case_tags(raw_arnaldo)
    assert_equal("multi_propriedades" in pattern_tags, True, "biblioteca tag multiplas propriedades")
    assert_equal("milho" in pattern_tags, True, "biblioteca tag milho")
    assert_equal("pecuaria_corte" in pattern_tags, True, "biblioteca tag pecuaria")
    pattern_selection = select_pattern_examples(raw_arnaldo)
    assert_equal(pattern_selection.examples[0].id, "arnaldo_melo_aprovado", "biblioteca seleciona exemplo arnaldo aprovado")
    assert_equal(any(example.has_expected for example in pattern_selection.examples), True, "biblioteca carrega saidas aprovadas")
    writer_prompt = build_writer_prompt(raw_arnaldo, technical_result.report_text, max_examples=2)
    assert_equal("GUIA DE ESTILO:" in writer_prompt, True, "prompt inclui guia")
    assert_equal("EXEMPLOS APROVADOS PARA IMITAR O PADRÃO:" in writer_prompt, True, "prompt inclui exemplos aprovados")
    assert_equal("DADOS BRUTOS DA VISITA:" in writer_prompt, True, "prompt inclui dados brutos")
    assert_equal("RASCUNHO LOCAL ESTRUTURADO:" in writer_prompt, True, "prompt inclui rascunho local")
    assert_equal(technical_result.notes.client, "Arnaldo Moreira", "redator tecnico cliente")
    assert_equal(len(technical_result.notes.properties), 4, "redator tecnico quantidade propriedades")
    assert_equal(len(technical_result.notes.equipment), 12, "redator tecnico quantidade maquinarios")
    assert_equal(technical_parsed.get("cliente"), "Arnaldo Moreira", "texto tecnico campo cliente")
    assert_equal(technical_parsed.get("cpf_cnpj"), "863.546.041-34", "texto tecnico campo cpf")
    assert_equal(technical_parsed.get("area_total_ha"), 271.04, "texto tecnico area total")
    assert_equal(technical_parsed.get("area_pastagens_ha"), 162.62, "texto tecnico area pastagens")
    assert_equal(technical_parsed.get("area_cultivo_ha"), 108.42, "texto tecnico area cultivo")
    assert_equal(technical_parsed.get("principais_culturas"), "Milho, Pastagens de Braquiarão", "texto tecnico culturas")
    assert_equal(len(technical_parsed.get("equipamentos", [])), 12, "texto tecnico equipamentos")
    technical_output = generate_report(technical_result.report_text, output_path=OUTPUT_DIR / "arnaldo-dados-brutos.xlsx")
    technical_worksheet = load_workbook(technical_output).active
    assert_equal(technical_worksheet["B4"].value, "Arnaldo Moreira", "texto tecnico cliente excel")
    assert_equal(technical_worksheet["A18"].value, "Fazenda Santa Rita", "texto tecnico primeira propriedade excel")
    assert_equal(technical_worksheet["A21"].value, "Fazenda Engenho de São Benedito", "texto tecnico quarta propriedade excel")
    assert_equal(technical_worksheet["D18"].value, 19.36, "texto tecnico area primeira propriedade excel")
    assert_equal(technical_worksheet["A46"].value, None, "texto tecnico limpa equipamento antigo")
    assert_equal(technical_worksheet["A51"].value, "Plantadeira Baldan 10 Linhas", "texto tecnico primeiro equipamento excel")
    assert_equal(technical_worksheet["A54"].value, "Trator New Holland 7630", "texto tecnico quarto equipamento excel")

    examples = {example.id: example for example in load_pattern_examples()}
    assert_equal("arnaldo_melo_aprovado" in examples, True, "biblioteca contem arnaldo aprovado")
    assert_equal("sandro_mabel_pecuaria_grande_escala" in examples, True, "biblioteca contem sandro aprovado")
    assert_equal(examples["arnaldo_melo_aprovado"].has_expected, True, "arnaldo aprovado tem resposta")
    assert_equal(examples["sandro_mabel_pecuaria_grande_escala"].has_expected, True, "sandro aprovado tem resposta")
    sandro_tags = classify_case_tags(examples["sandro_mabel_pecuaria_grande_escala"].raw_text)
    assert_equal("multi_propriedades" in sandro_tags, True, "sandro tag multiplas propriedades")
    assert_equal("grande_escala" in sandro_tags, True, "sandro tag grande escala")
    assert_equal("pecuaria_corte" in sandro_tags, True, "sandro tag pecuaria")
    sandro_result = generate_technical_report(examples["sandro_mabel_pecuaria_grande_escala"].raw_text)
    sandro_parsed = parse_report_data(sandro_result.report_text)
    assert_equal(len(sandro_result.notes.properties), 3, "redator sandro quantidade propriedades")
    assert_equal(sandro_parsed.get("area_total_ha"), 7681.08, "redator sandro area total")
    assert_equal(sandro_parsed.get("area_pastagens_ha"), 5376.76, "redator sandro area pastagens")
    assert_equal(sandro_parsed.get("area_cultivo_ha"), 2304.32, "redator sandro area cultivo")

    approved_pattern_cases = {
        "divino_piscicultura_pecuaria": {
            "tags": ("piscicultura", "area_alugada", "pecuaria_corte"),
            "properties": 2,
            "areas": (33.88, 23.72, 10.16),
        },
        "alfredo_ciclo_completo_confinamento": {
            "tags": ("ciclo_completo", "confinamento_seca", "piquetes"),
            "properties": 2,
            "areas": (532.40, 372.68, 159.72),
        },
        "socrates_pecuaria_duas_unidades": {
            "tags": ("poco_artesiano", "rotacionado", "multi_propriedades"),
            "properties": 2,
            "areas": (329.31, 230.53, 98.79),
        },
        "vanderlei_recria_projetos_futuros": {
            "tags": ("projetos_futuros", "reforma_pastagem", "aquisicao_animais"),
            "properties": 1,
            "areas": (94.38, 66.07, 28.31),
        },
    }
    for example_id, expectations in approved_pattern_cases.items():
        assert_equal(example_id in examples, True, f"biblioteca contem {example_id}")
        example = examples[example_id]
        assert_equal(example.has_expected, True, f"{example_id} tem resposta aprovada")
        assert_equal(bool(example.final_workbook) and Path(example.final_workbook).exists(), True, f"{example_id} tem planilha final")
        case_tags = classify_case_tags(example.raw_text)
        for tag in expectations["tags"]:
            assert_equal(tag in case_tags, True, f"{example_id} tag {tag}")
        case_selection = select_pattern_examples(example.raw_text)
        assert_equal(case_selection.examples[0].id, example_id, f"{example_id} seleciona exemplo correspondente")
        case_result = generate_technical_report(example.raw_text)
        case_parsed = parse_report_data(case_result.report_text)
        assert_equal(len(case_result.notes.properties), expectations["properties"], f"{example_id} quantidade propriedades")
        total, pasture, crop = expectations["areas"]
        assert_equal(case_parsed.get("area_total_ha"), total, f"{example_id} area total")
        assert_equal(case_parsed.get("area_pastagens_ha"), pasture, f"{example_id} area pastagens")
        assert_equal(case_parsed.get("area_cultivo_ha"), crop, f"{example_id} area cultivo")

    photo_output = generate_report(
        CASES[0]["source"].read_text(encoding="utf-8"),
        sorted((ROOT / "sample_photos").glob("*.jpg")),
        OUTPUT_DIR / "fotos-com-moldura.xlsx",
    )
    photo_worksheet = load_workbook(photo_output).active
    report_photo_rows = [
        int(image.anchor._from.row) + 1
        for image in photo_worksheet._images
        if getattr(getattr(image, "anchor", None), "_from", None) is not None and int(image.anchor._from.row) + 1 >= 200
    ]
    assert_equal(report_photo_rows, [213, 232, 251], "posicao das fotos inseridas")
    assert_equal(photo_worksheet["D212"].value, "Foto 01", "legenda primeira foto")
    assert_equal(photo_worksheet["D231"].value, "Foto 02", "legenda segunda foto")
    assert_equal(photo_worksheet["D212"].border.top.style, "medium", "borda superior primeira foto")
    assert_equal(photo_worksheet["D212"].border.left.style, "medium", "borda esquerda primeira foto")
    assert_equal(photo_worksheet["K229"].border.right.style, "medium", "borda direita primeira foto")
    assert_equal(photo_worksheet["K229"].border.bottom.style, "medium", "borda inferior primeira foto")
    prepared_photo = OUTPUT_DIR / "fotos-com-moldura-images" / "foto-01.jpg"
    with Image.open(prepared_photo) as image:
        assert_equal(image.size, (520, 330), "tamanho da foto preparada")
        left_edge = [image.getpixel((0, y)) for y in range(image.height)]
        right_edge = [image.getpixel((image.width - 1, y)) for y in range(image.height)]
        if all(pixel == (255, 255, 255) for pixel in left_edge + right_edge):
            raise AssertionError("foto preparada contem faixa branca lateral")

    coordinate_photo = OUTPUT_DIR / "foto-com-coordenada.jpg"
    marker = Image.new("RGB", (640, 480), "white")
    for x in range(40):
        for y in range(40):
            marker.putpixel((x, y), (255, 0, 0))
    marker.save(coordinate_photo)
    generate_report(
        CASES[0]["source"].read_text(encoding="utf-8"),
        [coordinate_photo],
        OUTPUT_DIR / "foto-coordenada-preservada.xlsx",
    )
    prepared_coordinate_photo = OUTPUT_DIR / "foto-coordenada-preservada-images" / "foto-01.jpg"
    with Image.open(prepared_coordinate_photo) as image:
        red, green, blue = image.getpixel((0, 0))
        if not (red > 150 and green < 90 and blue < 90):
            raise AssertionError("canto superior esquerdo da foto foi cortado")

    for case in CASES:
        text = case["source"].read_text(encoding="utf-8")
        parsed = parse_report_data(text)
        for field, expected in case["fields"].items():
            assert_equal(parsed.get(field), expected, f"{case['name']} campo {field}")

        output = generate_report(text, output_path=case["output"])
        worksheet = load_workbook(output).active
        for cell, expected in case["cells"].items():
            assert_equal(worksheet[cell].value, expected, f"{case['name']} célula {cell}")

        reviewed_fields = {field: str(parsed[field]) for field in case["fields"] if field.startswith("area_")}
        reviewed = parse_review_data(
            json.dumps({"parsed": parsed, "fields": reviewed_fields}, ensure_ascii=False),
            text,
        )
        reviewed_output = generate_report(reviewed, output_path=OUTPUT_DIR / f"{case['name']}-revisado.xlsx")
        reviewed_worksheet = load_workbook(reviewed_output).active
        for cell in ("D18", "E18", "F18"):
            assert_equal(reviewed_worksheet[cell].value, case["cells"][cell], f"{case['name']} revisão célula {cell}")

    print("OK: validações de extração e preenchimento concluídas.")


if __name__ == "__main__":
    main()
