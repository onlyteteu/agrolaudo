"""Teste offline da camada de fotos (apply_photos -> planilha).

Gera fotos sinteticas unicas, aplica no template e confere:
- legendas "Foto NN" sequenciais, sem numero pulado;
- cada imagem ancorada logo abaixo da propria legenda;
- linhas do slot somam altura suficiente para a imagem (sem foto
  invadindo a legenda do slot seguinte);
- prepare_photo preserva proporcao (letterbox, sem esticar).
"""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from PIL import Image

from relatorio_app.field_mapping import PHOTO_MAX_SIZE
from relatorio_app.report_engine import (
    DEFAULT_TEMPLATE,
    PHOTO_BORDER_COLOR,
    PHOTO_BORDER_PX,
    apply_photos,
    photo_anchor,
    prepare_photo,
)

OUTPUT_DIR = ROOT / "outputs" / "validation" / "photos"
# Cobre os slots estreitos do template (1, 2, 5, 8, 11, 12) E slots gerados
# alem dos 39 do modelo — nao pode haver limite de fotos.
PHOTO_COUNT = 45


def assert_true(condition: bool, label: str) -> None:
    if not condition:
        raise AssertionError(label)


def make_photos(directory: Path) -> list[Path]:
    directory.mkdir(parents=True, exist_ok=True)
    paths = []
    for index in range(PHOTO_COUNT):
        # tamanhos variados (retrato, paisagem) e cor unica por foto
        size = (400 + index * 7, 300) if index % 2 == 0 else (300, 400 + index * 7)
        image = Image.new("RGB", size, ((10 + index * 20) % 256, 120, 60))
        path = directory / f"origem-{index + 1:02d}.jpg"
        image.save(path)
        paths.append(path)
    return paths


def row_px(worksheet, row: int) -> float:
    dim = worksheet.row_dimensions.get(row)
    default_pt = worksheet.sheet_format.defaultRowHeight or 15
    pt = dim.height if dim is not None and dim.height is not None else default_pt
    return pt / 0.75


def main() -> None:
    photos = make_photos(OUTPUT_DIR / "origem")
    workbook = load_workbook(DEFAULT_TEMPLATE)
    worksheet = workbook.active
    apply_photos(worksheet, photos, OUTPUT_DIR / "imagens")

    added = [img for img in worksheet._images if isinstance(img.anchor, str)]
    images = sorted(added, key=lambda img: coordinate_to_tuple(img.anchor)[0])
    assert_true(len(images) == PHOTO_COUNT, f"esperava {PHOTO_COUNT} imagens, achou {len(images)}")

    for index, image in enumerate(images, start=1):
        from_row, from_col, to_row, _ = photo_anchor(index - 1)
        caption_row = from_row + 1
        anchor_row = coordinate_to_tuple(image.anchor)[0]  # 1-based

        caption = worksheet.cell(row=caption_row, column=from_col + 1).value
        assert_true(caption == f"Foto {index:02d}", f"slot {index}: legenda '{caption}'")
        assert_true(anchor_row == caption_row + 1, f"slot {index}: imagem ancorada em {anchor_row}, legenda em {caption_row}")

        slot_px = sum(row_px(worksheet, row) for row in range(caption_row + 1, to_row + 2))
        assert_true(
            slot_px >= PHOTO_MAX_SIZE[1],
            f"slot {index}: {slot_px:.0f}px disponiveis para imagem de {PHOTO_MAX_SIZE[1]}px (foto invade slot seguinte)",
        )

    # slots extras entram na area de impressao
    expected_bottom = photo_anchor(PHOTO_COUNT - 1)[2] + 2
    assert_true(
        worksheet.print_area is not None and worksheet.print_area.endswith(str(expected_bottom)),
        f"print_area '{worksheet.print_area}' deveria terminar na linha {expected_bottom}",
    )

    # proporcao: foto retrato deve virar letterbox, nao ser esticada
    tall = OUTPUT_DIR / "origem" / "retrato.jpg"
    Image.new("RGB", (200, 800), (200, 40, 40)).save(tall)
    prepared = prepare_photo(tall, OUTPUT_DIR / "imagens", 99)
    with Image.open(prepared) as framed:
        assert_true(framed.size == PHOTO_MAX_SIZE, f"tamanho final {framed.size}")
        bar_pixel = framed.getpixel((PHOTO_BORDER_PX + 40, PHOTO_MAX_SIZE[1] // 2))
        center_pixel = framed.getpixel((PHOTO_MAX_SIZE[0] // 2, PHOTO_MAX_SIZE[1] // 2))
        assert_true(bar_pixel == PHOTO_BORDER_COLOR, f"lateral deveria ser barra {PHOTO_BORDER_COLOR}, veio {bar_pixel} (foto esticada)")
        assert_true(center_pixel[0] > 150, f"centro deveria ser a foto, veio {center_pixel}")

    print("validate_photos: OK")


if __name__ == "__main__":
    main()
