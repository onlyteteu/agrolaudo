"""Teste offline do caminho estruturado (dict da IA -> planilha).

Simula a saida JSON do Gemini com um dicionario fixo, passa por
``coerce_structured_data`` e gera a planilha, conferindo celula a celula.
Nao depende de chave do Gemini: valida toda a montagem deterministica.
"""

from __future__ import annotations

from pathlib import Path
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from relatorio_app.report_engine import generate_report
from relatorio_app.report_schema import coerce_structured_data

OUTPUT_DIR = ROOT / "outputs" / "validation"

# Simula exatamente o que o Gemini deveria devolver no schema estruturado.
AI_JSON = {
    "cliente": "Joao da Silva",
    "cpf_cnpj": "123.456.789-00",
    "data_visita": "15/06/2026",
    "cidade_uf": "Morrinhos-GO",
    "imoveis": [
        {
            "nome": "Fazenda Boa Vista",
            "area_total_ha": 120.50,
            "area_pastagens_ha": 80.0,
            "area_cultivo_ha": 40.5,
            "atividade_principal": "Pecuaria de corte",
            "principais_culturas": "Pastagens de Braquiarao",
        },
        {
            "nome": "Fazenda Santa Luzia",
            "area_total_ha": 200.0,
            "area_pastagens_ha": 140.0,
            "area_cultivo_ha": 60.0,
            "atividade_principal": "Pecuaria de corte",
            "principais_culturas": "Pastagens de Mombaca",
        },
    ],
    "benfeitorias_descricao": "Na Fazenda Boa Vista, ha curral e cochos cobertos. Na Fazenda Santa Luzia, ha galpao de armazenagem.",
    "benfeitorias_conservacao": "BOM",
    "equipamentos": [
        {"descricao": "Trator Massey Ferguson 4292", "fabricante": "Massey Ferguson", "modelo": "4292"},
        {"descricao": "Grade aradora 16 discos"},
    ],
    "investimentos_comentarios": "Reforma de pastagens em andamento.",
    "insumos_comentarios": "PECUARIA DE CORTE EM 320,50 HA.",
    "outros_comentarios": "Operacao conduzida de forma integrada.",
    "conclusao": "Atividade regular e tecnicamente recomendada, condicionada a conferencia documental.",
}


def assert_equal(actual, expected, label: str) -> None:
    if isinstance(expected, float):
        if actual is None or round(float(actual), 2) != round(expected, 2):
            raise AssertionError(f"{label}: esperado {expected!r}, veio {actual!r}")
        return
    if actual != expected:
        raise AssertionError(f"{label}: esperado {expected!r}, veio {actual!r}")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    data = coerce_structured_data(AI_JSON)

    # Marca de origem estruturada e consolidacao de areas.
    assert_equal(data.get("_structured"), True, "marca _structured")
    assert_equal(data.get("imovel_nome"), "Fazenda Boa Vista; Fazenda Santa Luzia", "nomes consolidados")
    assert_equal(data.get("area_total_ha"), 320.5, "area total consolidada")
    assert_equal(data.get("area_pastagens_ha"), 220.0, "area pastagens consolidada")
    assert_equal(data.get("area_cultivo_ha"), 100.5, "area cultivo consolidada")
    assert_equal(len(data.get("equipamentos", [])), 2, "quantidade de equipamentos")

    output = generate_report(data, output_path=OUTPUT_DIR / "estruturado.xlsx")
    ws = load_workbook(output).active

    assert_equal(ws["B4"].value, "JOAO DA SILVA", "cliente em B4")
    assert_equal(ws["B5"].value, "123.456.789-00", "cpf em B5")
    assert_equal(ws["A18"].value, "FAZENDA BOA VISTA", "propriedade 1 em A18")
    assert_equal(ws["D18"].value, 120.5, "area total propriedade 1 em D18")
    assert_equal(ws["E18"].value, 80.0, "pastagens propriedade 1 em E18")
    assert_equal(ws["F18"].value, 40.5, "cultivo propriedade 1 em F18")
    assert_equal(ws["A19"].value, "FAZENDA SANTA LUZIA", "propriedade 2 em A19")
    assert_equal(ws["D19"].value, 200.0, "area total propriedade 2 em D19")
    assert_equal(ws["A27"].value.startswith("Na Fazenda Boa Vista"), True, "benfeitorias bloco 1 em A27")
    assert_equal(ws["A30"].value.startswith("Na Fazenda Santa Luzia"), True, "benfeitorias bloco 2 em A30")
    # O template traz as linhas 27-29 ocultas; o laudo gerado nao pode
    # esconder o primeiro bloco de benfeitorias.
    for row in range(27, 34):
        dim = ws.row_dimensions.get(row)
        hidden = bool(dim.hidden) if dim is not None else False
        assert_equal(hidden, False, f"linha {row} de benfeitorias visivel")
    assert_equal(ws["A51"].value, "Trator Massey Ferguson 4292", "equipamento 1 em A51")
    assert_equal(ws["A52"].value, "Grade aradora 16 discos", "equipamento 2 em A52")
    assert_equal(ws["J4"].value, "Data da Visita:15/06/2026", "data da visita em J4")
    assert_equal(ws["B193"].value.startswith("Atividade regular"), True, "conclusao em B193")

    # Com 8 imoveis ha insercao de linhas (propriedades E benfeitorias); as
    # flags de linha oculta do template precisam acompanhar o deslocamento,
    # senao escondem equipamentos/benfeitorias (insert_rows nao move
    # row_dimensions). Invariante: linha com conteudo nunca fica oculta.
    big = coerce_structured_data(
        {
            "cliente": "Teste Offset",
            "imoveis": [{"nome": f"Fazenda {i}", "area_total_ha": 10.0 + i} for i in range(8)],
            "benfeitorias_descricao": " ".join(f"Na Fazenda {i}, ha curral e cochos." for i in range(8)),
            "benfeitorias_conservacao": "BOM",
            "equipamentos": [{"descricao": "Trator Massey Ferguson 4292"}],
        }
    )
    big_output = generate_report(big, output_path=OUTPUT_DIR / "estruturado-8-imoveis.xlsx")
    big_ws = load_workbook(big_output).active
    for row in range(8, 201):  # linha 7 e campo opcional oculto de proposito no modelo
        dim = big_ws.row_dimensions.get(row)
        if dim is None or not dim.hidden:
            continue
        content = [cell.value for cell in big_ws[row][:11] if cell.value not in (None, "")]
        assert_equal(content, [], f"linha {row} oculta com conteudo")

    print("OK: caminho estruturado (dict da IA -> planilha) validado.")


if __name__ == "__main__":
    main()
